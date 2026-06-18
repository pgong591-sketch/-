"""Build aligned management/non-subject cost summaries from source workbooks.

This script handles three 202603 source files:
1) management center dept income/cost workbook
2) non-subject management center dept income/cost workbook
3) non-subject teaching fee workbook (as campus-class income source)

Outputs one Excel workbook with multiple sheets under data/output.
"""

from __future__ import annotations

import argparse
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, Iterable, List, Sequence

import numpy as np
import pandas as pd
from openpyxl import load_workbook


BASE_DIR = Path(__file__).resolve().parent.parent
INCOMING_DIR = BASE_DIR / "data" / "incoming"
TEMPLATE_DIR = BASE_DIR / "data" / "templates"
OUTPUT_DIR = BASE_DIR / "data" / "output"


MGMT_SOURCE = INCOMING_DIR / "202603_mgmt_dept_income_cost.xlsx"
NON_SUBJECT_SOURCE = INCOMING_DIR / "202603_non_subject_mgmt_dept_income_cost.xlsx"
TEACHING_FEE_SOURCE = INCOMING_DIR / "202603_non_subject_teaching_fee.xlsx"


def clean_text(value: object) -> str:
    if value is None:
        return ""
    text = str(value)
    text = text.replace("\u3000", " ").strip()
    return re.sub(r"\s+", "", text)


def parse_dept_header(value: object) -> str:
    text = clean_text(value)
    if not text:
        return ""
    match = re.search(r"部门档案[:：](.+)$", text)
    if match:
        dept = clean_text(match.group(1))
        return dept.replace("】", "").replace("]", "")
    text = text.replace("【", "").replace("】", "")
    text = text.replace("部门档案:", "").replace("部门档案：", "")
    return clean_text(text)


def to_float(value: object) -> float:
    if value is None:
        return 0.0
    if isinstance(value, (int, float, np.floating)):
        if pd.isna(value):
            return 0.0
        return float(value)
    text = str(value).strip()
    if not text:
        return 0.0
    text = text.replace(",", "")
    try:
        return float(text)
    except ValueError:
        return 0.0


def excel_row_to_template_row(source_row_idx_0_based: int) -> int:
    """Source sheets have one extra header row versus template detail rows."""
    return source_row_idx_0_based - 1


@dataclass
class MatrixData:
    source_path: Path
    dept_names: List[str]
    detail: pd.DataFrame


def load_matrix_source(path: Path) -> MatrixData:
    raw = pd.read_excel(path, sheet_name=0, header=None)
    dept_names = [parse_dept_header(v) for v in raw.iloc[5, 3:].tolist()]
    dept_names = [d for d in dept_names if d]

    records: List[Dict[str, object]] = []
    started = False
    for idx in range(len(raw)):
        name = clean_text(raw.iat[idx, 0])
        if idx < 7:
            continue
        if name.startswith("单位:"):
            break
        if not name or name == "会计科目":
            continue
        started = True
        if name == "合计":
            continue

        rec: Dict[str, object] = {
            "source_row_idx": idx,
            "template_row_idx": excel_row_to_template_row(idx),
            "account_name": name,
        }
        for offset, dept in enumerate(dept_names):
            rec[dept] = to_float(raw.iat[idx, 3 + offset])
        records.append(rec)

    if not started:
        raise ValueError(f"未能从文件读取到明细数据: {path}")

    detail = pd.DataFrame(records)
    return MatrixData(source_path=path, dept_names=dept_names, detail=detail)


def find_template_path() -> Path:
    candidates = sorted(TEMPLATE_DIR.glob("*.xlsx"))
    if not candidates:
        raise FileNotFoundError(f"未找到模板文件: {TEMPLATE_DIR}")
    return candidates[0]


def parse_formula_refs(formula: object) -> List[int]:
    if not isinstance(formula, str):
        return []
    refs = [int(m) for m in re.findall(r"[A-Z]+(\d+)", formula)]
    # Keep detail-area rows only; ignore summary self-references.
    return [r for r in refs if 1 <= r <= 90]


@dataclass
class TemplateRules:
    mgmt_category_refs: Dict[str, List[int]]
    non_subject_category_refs: Dict[str, List[int]]
    mgmt_income_refs: List[int]
    mgmt_cost_refs: List[int]
    non_subject_income_refs: List[int]
    non_subject_cost_refs: List[int]


def load_template_rules(template_path: Path) -> TemplateRules:
    wb = load_workbook(template_path, data_only=False)
    mgmt_ws = wb.worksheets[6]  # 管理公司收入成本费用表
    non_ws = wb.worksheets[7]   # 非学科管理中心收入成本费用表

    def extract_rules(ws, rows: Sequence[int]) -> Dict[str, List[int]]:
        result: Dict[str, List[int]] = {}
        for row in rows:
            label = clean_text(ws.cell(row, 2).value)
            if not label:
                continue
            formula = ws.cell(row, 3).value
            refs = parse_formula_refs(formula)
            if not refs:
                refs = parse_formula_refs(ws.cell(row, 4).value)
            result[label] = refs
        return result

    mgmt_category_refs = extract_rules(mgmt_ws, range(94, 104))
    non_subject_category_refs = extract_rules(non_ws, range(92, 102))

    return TemplateRules(
        mgmt_category_refs=mgmt_category_refs,
        non_subject_category_refs=non_subject_category_refs,
        mgmt_income_refs=[6, 7, 8],   # no investment income
        mgmt_cost_refs=[9, 39, 59],
        non_subject_income_refs=[6],
        non_subject_cost_refs=[7, 37, 57],
    )


def sum_rows_by_refs(detail: pd.DataFrame, dept_names: Sequence[str], refs: Iterable[int]) -> Dict[str, float]:
    ref_set = set(int(r) for r in refs)
    if not ref_set:
        return {d: 0.0 for d in dept_names}
    selected = detail[detail["template_row_idx"].isin(ref_set)]
    return {d: float(selected[d].sum()) for d in dept_names}


def project_row(
    raw_row: Dict[str, float],
    raw_to_target: Dict[str, str],
    target_cols: Sequence[str],
) -> Dict[str, float]:
    projected = {c: 0.0 for c in target_cols}
    for src_dept, amount in raw_row.items():
        target = raw_to_target.get(src_dept)
        if not target:
            continue
        if target not in projected:
            continue
        projected[target] += amount
    return projected


def build_summary_table(
    detail: pd.DataFrame,
    dept_names: Sequence[str],
    category_refs: Dict[str, List[int]],
    income_refs: Sequence[int],
    cost_refs: Sequence[int],
    display_map: Dict[str, str],
    display_cols: Sequence[str],
    extra_income_row: Dict[str, float] | None = None,
) -> pd.DataFrame:
    base_order = [
        "学生福利及教具",
        "房租水电",
        "人工",
        "税金",
        "销售费用",
        "办公",
        "差旅交际费",
        "折旧及摊销",
        "管理中心",
        "财务费用",
        "其他",
    ]

    rows: List[Dict[str, object]] = []
    cat_values: Dict[str, Dict[str, float]] = {}

    for label in base_order:
        refs = category_refs.get(label, [])
        raw_amounts = sum_rows_by_refs(detail, dept_names, refs)
        projected = project_row(raw_amounts, display_map, display_cols)
        cat_values[label] = projected

    income_projected = project_row(
        sum_rows_by_refs(detail, dept_names, income_refs),
        display_map,
        display_cols,
    )
    cost_projected = project_row(
        sum_rows_by_refs(detail, dept_names, cost_refs),
        display_map,
        display_cols,
    )

    # 其他 = 成本费用合计 - 已归类项(含财务费用、管理中心等)
    subtotal_no_other = {c: 0.0 for c in display_cols}
    for label in base_order:
        if label == "其他":
            continue
        for col in display_cols:
            subtotal_no_other[col] += cat_values[label][col]
    cat_values["其他"] = {
        col: cost_projected[col] - subtotal_no_other[col]
        for col in display_cols
    }

    for label in base_order:
        row = {"费用": label}
        row.update(cat_values[label])
        row["合计"] = float(sum(row[c] for c in display_cols))
        rows.append(row)

    dep_row = next(r for r in rows if r["费用"] == "折旧及摊销")
    dep_values = {c: float(dep_row[c]) for c in display_cols}

    def append_metric(label: str, values: Dict[str, float]) -> None:
        row = {"费用": label}
        row.update(values)
        row["合计"] = float(sum(values[c] for c in display_cols))
        rows.append(row)

    append_metric("成本费用合计", cost_projected)
    append_metric("收入合计", income_projected)

    if extra_income_row is not None:
        extra_values = {c: float(extra_income_row.get(c, 0.0)) for c in display_cols}
        append_metric("其中：校区上课收入", extra_values)

    append_metric(
        "净利润",
        {c: income_projected[c] - cost_projected[c] for c in display_cols},
    )
    append_metric("折旧及摊销(指标)", dep_values)
    append_metric(
        "实际成本",
        {c: cost_projected[c] - dep_values[c] for c in display_cols},
    )
    append_metric(
        "净利润（不含折旧与摊销）",
        {c: income_projected[c] - (cost_projected[c] - dep_values[c]) for c in display_cols},
    )

    out = pd.DataFrame(rows)
    cols = ["费用", "合计", *display_cols]
    out = out[cols]
    return out


def teaching_fee_summary(path: Path) -> pd.DataFrame:
    df = pd.read_excel(path, sheet_name=0)
    # 口径使用打印部门名称。
    grp = (
        df.groupby("打印部门名称", dropna=False)["总收入"]
        .sum()
        .reset_index()
        .rename(columns={"打印部门名称": "source_dept", "总收入": "amount"})
    )
    grp["source_dept"] = grp["source_dept"].apply(clean_text)
    grp["amount"] = grp["amount"].apply(to_float)
    return grp


def build_teaching_income_projection(grp: pd.DataFrame) -> Dict[str, float]:
    result = {
        "执行委员会": 0.0,
        "品牌部": 0.0,
        "财务法务部": 0.0,
        "教研部": 0.0,
        "行政教务部": 0.0,
        "总经办": 0.0,
        "人力服务部": 0.0,
        "幼儿园管理中心部": 0.0,
        "咨询部": 0.0,
        "学术中心": 0.0,
        "信息部": 0.0,
    }
    source_to_target = {
        "素质管理中心": "执行委员会",
        "素质中心教研部": "教研部",
        "学术中心": "学术中心",
    }
    for _, row in grp.iterrows():
        target = source_to_target.get(clean_text(row["source_dept"]))
        if not target:
            continue
        result[target] += to_float(row["amount"])
    return result


def merged_template_table(
    mgmt_table: pd.DataFrame,
    non_table: pd.DataFrame,
) -> pd.DataFrame:
    col_order = [
        "董事会",
        "品牌部",
        "财务法务部",
        "教研部",
        "行政教务部",
        "新项目拓展部",
        "总经办",
        "人力服务部",
        "幼儿园管理中心部",
        "董事助理办公室",
        "信息部",
        "未对齐_咨询部",
        "未对齐_学术中心",
    ]

    mgmt_map = {
        "总经办": "董事会",            # 管理中心显示列“总经办”实为董事会口径
        "品牌部": "品牌部",
        "财务部": "财务法务部",
        "咨询部": "教研部",            # 模板合并区按教研部列叠加
        "行政教务部": "行政教务部",
        "新项目拓展部": "新项目拓展部",
        "集团办": "总经办",            # 集团办与总经办对齐
        "人力服务部": "人力服务部",
        "幼儿园管理中心部": "幼儿园管理中心部",
        "董事助理办公室": "董事助理办公室",
        "信息部": "信息部",
    }
    non_map = {
        "执行委员会": "董事会",
        "品牌部": "品牌部",
        "财务法务部": "财务法务部",
        "教研部": "教研部",
        "行政教务部": "行政教务部",
        "总经办": "总经办",
        "人力服务部": "人力服务部",
        "幼儿园管理中心部": "幼儿园管理中心部",
        "咨询部": "未对齐_咨询部",      # 运营部映射到咨询部展示列，不强行并入主口径
        "学术中心": "未对齐_学术中心",
        "信息部": "信息部",
    }

    row_order = [
        "学生福利及教具",
        "房租水电",
        "人工",
        "税金",
        "销售费用",
        "办公",
        "差旅交际费",
        "折旧及摊销",
        "管理中心",
        "财务费用",
        "其他",
    ]
    add_non_rows = {"房租水电", "人工", "销售费用", "办公", "差旅交际费", "其他"}

    mgmt_idx = mgmt_table.set_index("费用")
    non_idx = non_table.set_index("费用")

    out_rows: List[Dict[str, object]] = []

    for label in row_order:
        row = {"费用": label}
        for col in col_order:
            row[col] = 0.0

        if label in mgmt_idx.index:
            mvals = mgmt_idx.loc[label].to_dict()
            for src, dst in mgmt_map.items():
                row[dst] += to_float(mvals.get(src, 0.0))

        if label in add_non_rows and label in non_idx.index:
            nvals = non_idx.loc[label].to_dict()
            for src, dst in non_map.items():
                row[dst] += to_float(nvals.get(src, 0.0))

        row["合计"] = float(sum(row[c] for c in col_order))
        out_rows.append(row)

    total_row = {"费用": "成本费用合计"}
    for col in col_order:
        total_row[col] = float(sum(r[col] for r in out_rows))
    total_row["合计"] = float(sum(total_row[c] for c in col_order))
    out_rows.append(total_row)

    out = pd.DataFrame(out_rows)
    out = out[["费用", "合计", *col_order]]
    return out


def period_labels(period: str) -> Dict[str, str]:
    year = period[:4]
    month = int(period[4:6])
    quarter = (month - 1) // 3 + 1
    return {
        "period": period,
        "business_period": f"{year}-{period[4:6]}",
        "calendar_quarter": f"{year}Q{quarter}",
        "source_quarter_label": f"{year}年{quarter}季度",
    }


def tidy_numeric(df: pd.DataFrame, decimals: int = 2) -> pd.DataFrame:
    out = df.copy()
    for col in out.columns:
        if col in {"费用", "key", "value", "source_dept", "target_column"}:
            continue
        if pd.api.types.is_numeric_dtype(out[col]):
            vals = out[col].astype(float)
            vals = vals.where(~np.isclose(vals, 0.0, atol=1e-8), 0.0)
            out[col] = vals.round(decimals)
    return out


def build_report(output_path: Path, period: str) -> None:
    template_path = find_template_path()
    rules = load_template_rules(template_path)

    mgmt = load_matrix_source(MGMT_SOURCE)
    non_subject = load_matrix_source(NON_SUBJECT_SOURCE)
    teaching_grp = teaching_fee_summary(TEACHING_FEE_SOURCE)

    mgmt_display_cols = [
        "总经办",
        "品牌部",
        "财务部",
        "咨询部",
        "行政教务部",
        "新项目拓展部",
        "集团办",
        "人力服务部",
        "幼儿园管理中心部",
        "董事助理办公室",
        "信息部",
    ]
    mgmt_map = {
        "董事会": "总经办",
        "品牌部": "品牌部",
        "财务法务部": "财务部",
        "咨询部": "咨询部",
        "行政综合部": "行政教务部",
        "新项目拓展部": "新项目拓展部",
        "总经办": "集团办",
        "人力服务部": "人力服务部",
        "幼儿园管理中心部": "幼儿园管理中心部",
        "董事助理办公室": "董事助理办公室",
        "信息部": "信息部",
    }

    non_display_cols = [
        "执行委员会",
        "品牌部",
        "财务法务部",
        "教研部",
        "行政教务部",
        "总经办",
        "人力服务部",
        "幼儿园管理中心部",
        "咨询部",
        "学术中心",
        "信息部",
    ]
    non_map = {
        "素质管理中心": "执行委员会",
        "品牌部": "品牌部",
        "财务法务部": "财务法务部",
        "素质中心教研部": "教研部",
        "行政综合部": "行政教务部",
        "总经办": "总经办",
        "人力服务部": "人力服务部",
        "运营部": "咨询部",
        "学术中心": "学术中心",
        "幼儿园管理中心部": "幼儿园管理中心部",
        "信息部": "信息部",
    }

    teaching_income_row = build_teaching_income_projection(teaching_grp)

    mgmt_summary = build_summary_table(
        detail=mgmt.detail,
        dept_names=mgmt.dept_names,
        category_refs=rules.mgmt_category_refs,
        income_refs=rules.mgmt_income_refs,
        cost_refs=rules.mgmt_cost_refs,
        display_map=mgmt_map,
        display_cols=mgmt_display_cols,
    )

    non_summary = build_summary_table(
        detail=non_subject.detail,
        dept_names=non_subject.dept_names,
        category_refs=rules.non_subject_category_refs,
        income_refs=rules.non_subject_income_refs,
        cost_refs=rules.non_subject_cost_refs,
        display_map=non_map,
        display_cols=non_display_cols,
        extra_income_row=teaching_income_row,
    )

    merged_summary = merged_template_table(mgmt_summary, non_summary)

    # Build fee detail view with mapped column tags.
    fee_tag_map = {
        "素质管理中心": "执行委员会",
        "素质中心教研部": "教研部",
        "学术中心": "学术中心",
    }
    teaching_detail = teaching_grp.copy()
    teaching_detail["target_column"] = teaching_detail["source_dept"].map(fee_tag_map).fillna("未映射")
    teaching_detail = teaching_detail.sort_values(["target_column", "source_dept"]).reset_index(drop=True)

    labels = period_labels(period)
    metadata = pd.DataFrame(
        [
            {"key": "period", "value": labels["period"]},
            {"key": "business_period", "value": labels["business_period"]},
            {"key": "calendar_quarter", "value": labels["calendar_quarter"]},
            {"key": "source_quarter_label", "value": labels["source_quarter_label"]},
            {"key": "rule_note_1", "value": "董事会与执行委员会同口径对齐"},
            {"key": "rule_note_2", "value": "集团办与总经办同口径对齐"},
            {"key": "rule_note_3", "value": "信息部已预留列，当前无源数据"},
            {"key": "rule_note_4", "value": "管理中心收入合计不体现投资收益（按主业/其他/营业外）"},
            {"key": "rule_note_5", "value": "非学科课酬用于“其中：校区上课收入”口径"},
        ]
    )

    mgmt_summary = tidy_numeric(mgmt_summary)
    non_summary = tidy_numeric(non_summary)
    merged_summary = tidy_numeric(merged_summary)
    teaching_detail = tidy_numeric(teaching_detail)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        metadata.to_excel(writer, sheet_name="元数据", index=False)
        mgmt_summary.to_excel(writer, sheet_name="管理中心归纳", index=False)
        non_summary.to_excel(writer, sheet_name="非学科管理中心归纳", index=False)
        merged_summary.to_excel(writer, sheet_name="管理中心成本分析_模板口径", index=False)
        teaching_detail.to_excel(writer, sheet_name="非学科课酬_校区上课收入", index=False)

    # Console validation snapshot
    mgmt_income = float(mgmt_summary.loc[mgmt_summary["费用"] == "收入合计", "合计"].iloc[0])
    non_income = float(non_summary.loc[non_summary["费用"] == "收入合计", "合计"].iloc[0])
    fee_income = float(non_summary.loc[non_summary["费用"] == "其中：校区上课收入", "合计"].iloc[0])
    merged_cost = float(merged_summary.loc[merged_summary["费用"] == "成本费用合计", "合计"].iloc[0])

    print(f"输出文件: {output_path}")
    print(f"管理中心收入合计(不含投资收益): {mgmt_income:.2f}")
    print(f"非学科收入合计: {non_income:.2f}")
    print(f"其中：校区上课收入: {fee_income:.2f}")
    print(f"管理中心成本分析(模板口径)成本费用合计: {merged_cost:.2f}")


def main() -> None:
    parser = argparse.ArgumentParser(description="Build management/non-subject summary workbook.")
    parser.add_argument(
        "--period",
        default="202603",
        help="Reporting period in YYYYMM format.",
    )
    parser.add_argument(
        "--output",
        default=str(OUTPUT_DIR / "202603_management_non_subject_alignment.xlsx"),
        help="Output xlsx path.",
    )
    args = parser.parse_args()

    build_report(output_path=Path(args.output), period=args.period)


if __name__ == "__main__":
    main()
