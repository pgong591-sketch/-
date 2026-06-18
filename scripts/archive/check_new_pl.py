"""查看损益表模板"""
import openpyxl

wb = openpyxl.load_workbook('data/损益表.xlsx', data_only=True)
if wb is None:
    # 附件文件可能在临时位置
    print("文件不在 data/ 目录")
else:
    ws = wb.worksheets[0]
    print(f'工作表: "{ws.title}" ({ws.max_row}行 x {ws.max_column}列)')
    for r in range(1, min((ws.max_row or 0) + 1, 35)):
        vals = []
        for c in range(1, min((ws.max_column or 0) + 1, 12)):
            v = ws.cell(r, c).value
            if v is None:
                vals.append('')
            else:
                s = str(v).strip()[:30]
                vals.append(s)
        print(f'  R{r}: {vals}')
