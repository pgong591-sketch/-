"""查看损益表模板结构"""
import openpyxl

wb = openpyxl.load_workbook('data/202603拔创中心损益表.xlsx', data_only=True)
ws = wb.worksheets[0]
print(f'工作表: "{ws.title}" ({ws.max_row}行 x {ws.max_column}列)')
for r in range(1, min((ws.max_row or 0) + 1, 30)):
    vals = []
    for c in range(1, min((ws.max_column or 0) + 1, 12)):
        v = ws.cell(r, c).value
        if v is None:
            vals.append('')
        else:
            s = str(v).strip()[:28]
            vals.append(s)
    print(f'  R{r}: {vals}')
