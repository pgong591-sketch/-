"""检查模板表头列"""
import openpyxl

wb = openpyxl.load_workbook('data/损益表.xlsx', data_only=True)
ws = wb.worksheets[0]
print(f'最大列: {ws.max_column}')

# 读表头行 (R4)
r = 4
for c in range(1, (ws.max_column or 0) + 1):
    v = ws.cell(r, c).value
    if v is not None:
        s = str(v).strip()[:25]
        print(f'  Col {c}: {s}')
