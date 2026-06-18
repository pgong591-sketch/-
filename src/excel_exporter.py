"""
Excel 导出模块

将查询结果 DataFrame 输出为格式化的 Excel 文件，支持：
- 自动设置列宽
- 表头样式（字体、背景色、边框）
- 数字格式（千分位、小数位）
- 冻结首行
- 多sheet导出
"""

from pathlib import Path
from typing import Optional, List, Dict, Any, Union
from datetime import datetime

import pandas as pd
import numpy as np


# ============================================================================
# 样式常量
# ============================================================================

# 默认 Excel 文件路径
DEFAULT_OUTPUT_DIR = Path(__file__).resolve().parent.parent / "data" / "output"

# 表头样式
HEADER_STYLE = {
    "font": {"bold": True, "size": 11, "color": "FFFFFF"},
    "fill": {"patternType": "solid", "fgColor": "4472C4"},
    "alignment": {"horizontal": "center", "vertical": "center", "wrap_text": True},
    "border": {"top": "thin", "bottom": "thin", "left": "thin", "right": "thin"},
}

# 数据行样式
DATA_STYLE = {
    "font": {"size": 10},
    "alignment": {"vertical": "center"},
    "border": {"top": "thin", "bottom": "thin", "left": "thin", "right": "thin"},
}

# 数字格式
NUMBER_FORMAT = "#,##0.00"
INTEGER_FORMAT = "#,##0"
PERCENTAGE_FORMAT = "0.00%"

# 报表标题样式
TITLE_STYLE = {
    "font": {"bold": True, "size": 16, "color": "1F4E79"},
    "alignment": {"horizontal": "center", "vertical": "center"},
}

# 小计行样式
SUBTOTAL_STYLE = {
    "font": {"bold": True, "size": 10},
    "fill": {"patternType": "solid", "fgColor": "D6E4F0"},
    "border": {"top": "medium", "bottom": "medium", "left": "thin", "right": "thin"},
}


# ============================================================================
# 核心导出函数
# ============================================================================

def export_to_excel(
    df: pd.DataFrame,
    file_path: Optional[Union[str, Path]] = None,
    sheet_name: str = "Sheet1",
    title: Optional[str] = None,
    subtitle: Optional[str] = None,
    number_format: str = NUMBER_FORMAT,
    column_widths: Optional[Dict[str, int]] = None,
    freeze_panes: bool = True,
    auto_filter: bool = True,
) -> str:
    """
    将 DataFrame 导出为格式化的 Excel 文件

    Args:
        df: 数据 DataFrame
        file_path: 输出文件路径（默认自动生成）
        sheet_name: 工作表名称
        title: 报表标题
        subtitle: 报表副标题
        number_format: 数字格式
        column_widths: 列宽映射 {列名: 宽度}
        freeze_panes: 是否冻结首行
        auto_filter: 是否启用自动筛选

    Returns:
        输出文件路径
    """
    if file_path is None:
        DEFAULT_OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        file_path = DEFAULT_OUTPUT_DIR / f"报表导出_{timestamp}.xlsx"

    file_path = Path(file_path)
    file_path.parent.mkdir(parents=True, exist_ok=True)

    with pd.ExcelWriter(file_path, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name=sheet_name, index=False, startrow=0)

        workbook = writer.book
        worksheet = writer.sheets[sheet_name]

        # 如果存在标题，在首行插入标题
        if title:
            # 需要在 DataFrame 上方插入标题行
            # 重新写入
            _write_with_title(writer, workbook, worksheet, df, sheet_name,
                              title, subtitle, number_format, column_widths,
                              freeze_panes, auto_filter)
        else:
            # 格式化
            _format_sheet(worksheet, workbook, df, sheet_name,
                          number_format, column_widths, freeze_panes, auto_filter)

    return str(file_path)


def _format_period_text(period: str) -> str:
    period = str(period or "").strip()
    if len(period) == 6 and period.isdigit():
        return f"{period[:4]}年{period[4:6]}月"
    return period


def _default_report_path(report_name: str) -> Path:
    DEFAULT_OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    return DEFAULT_OUTPUT_DIR / f"{report_name}_{timestamp}.xlsx"


def _normalize_export_columns(df: pd.DataFrame, preferred_columns: List[str]) -> pd.DataFrame:
    cols = []
    seen = set()
    for col in preferred_columns:
        if col in df.columns and col not in seen:
            cols.append(col)
            seen.add(col)
    if not cols:
        return df.copy()
    return df[cols].copy()


def _is_subtotal_text(value: object) -> bool:
    text = str(value or "")
    return any(token in text for token in ["合计", "小计", "总计", "净额", "流动资产：", "流动负债：", "非流动资产：", "非流动负债："])


def _coerce_excel_number(value: object, is_percent: bool = False) -> Optional[float]:
    if value is None or pd.isna(value):
        return None
    if isinstance(value, (np.integer, int, np.floating, float)):
        number = float(value)
        return number / 100 if is_percent and abs(number) > 1 else number
    text = str(value).strip().replace(",", "")
    if not text:
        return None
    percent = text.endswith("%")
    if percent:
        text = text[:-1]
    try:
        number = float(text)
    except ValueError:
        return None
    if percent or is_percent:
        return number / 100 if abs(number) > 1 else number
    return number


def _apply_statement_style(file_path: Union[str, Path], sheet_name: str, header_row: int = 3) -> None:
    from openpyxl import load_workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

    wb = load_workbook(file_path)
    ws = wb[sheet_name]

    thin = Side(style="thin", color="B7C4D6")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill("solid", fgColor="2F5597")
    subtotal_fill = PatternFill("solid", fgColor="D9EAF7")
    section_fill = PatternFill("solid", fgColor="EAF2F8")
    amount_headers = {"期末余额", "年初余额", "期末余额2", "年初余额2", "本期金额", "累计金额", "金额"}

    ws.sheet_view.showGridLines = False
    ws.row_dimensions[1].height = 24
    ws.row_dimensions[header_row].height = 22

    headers = {ws.cell(header_row, col).value: col for col in range(1, ws.max_column + 1)}
    amount_cols = {idx for name, idx in headers.items() if name in amount_headers or "金额" in str(name or "") or "余额" in str(name or "")}

    for cell in ws[header_row]:
        cell.font = Font(name="微软雅黑", size=10, bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    for row in range(header_row + 1, ws.max_row + 1):
        values = [ws.cell(row, col).value for col in range(1, ws.max_column + 1)]
        text = " ".join(str(v or "") for v in values)
        is_subtotal = "是" in text or any(_is_subtotal_text(v) for v in values)
        is_section = "：" in text and not is_subtotal

        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row, col)
            cell.font = Font(name="微软雅黑", size=10, bold=is_subtotal or is_section)
            cell.border = border
            cell.alignment = Alignment(
                horizontal="right" if col in amount_cols else "left",
                vertical="center",
                wrap_text=True,
            )
            if col in amount_cols and cell.value is not None:
                cell.number_format = NUMBER_FORMAT
            if is_subtotal:
                cell.fill = subtotal_fill
            elif is_section:
                cell.fill = section_fill

    ws.freeze_panes = f"A{header_row + 1}"
    wb.save(file_path)


def _write_with_title(
    writer, workbook, worksheet, df, sheet_name,
    title, subtitle, number_format, column_widths,
    freeze_panes, auto_filter,
):
    """带标题的写入"""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    # 清除原有内容
    for row in worksheet.iter_rows():
        for cell in row:
            cell.value = None

    # 写入标题
    title_row = 1
    subtitle_row = 2
    header_row = 3 if subtitle else 2
    data_start_row = header_row + 1

    # 标题
    title_font = Font(bold=True, size=16, color="1F4E79")
    title_align = Alignment(horizontal="center", vertical="center")
    worksheet.merge_cells(
        start_row=title_row, start_column=1,
        end_row=title_row, end_column=len(df.columns)
    )
    title_cell = worksheet.cell(row=title_row, column=1, value=title)
    title_cell.font = title_font
    title_cell.alignment = title_align

    # 副标题
    row_offset = title_row
    if subtitle:
        row_offset = subtitle_row
        subtitle_font = Font(size=10, color="666666", italic=True)
        subtitle_align = Alignment(horizontal="center", vertical="center")
        worksheet.merge_cells(
            start_row=subtitle_row, start_column=1,
            end_row=subtitle_row, end_column=len(df.columns)
        )
        st_cell = worksheet.cell(row=subtitle_row, column=1, value=subtitle)
        st_cell.font = subtitle_font
        st_cell.alignment = subtitle_align

    # 写入列名
    header_row_num = header_row
    for col_idx, col_name in enumerate(df.columns, 1):
        cell = worksheet.cell(row=header_row_num, column=col_idx, value=col_name)

    # 写入数据
    for row_idx, (_, row_data) in enumerate(df.iterrows(), data_start_row):
        for col_idx, col_name in enumerate(df.columns, 1):
            cell = worksheet.cell(row=row_idx, column=col_idx)
            val = row_data[col_name]
            if pd.isna(val):
                cell.value = None
            elif isinstance(val, (np.integer,)):
                cell.value = int(val)
            elif isinstance(val, (np.floating, float)):
                cell.value = float(val)
            elif isinstance(val, np.bool_):
                cell.value = bool(val)
            else:
                cell.value = val

    _format_sheet(worksheet, workbook, df, sheet_name,
                  number_format, column_widths, freeze_panes, auto_filter,
                  header_row=header_row_num, data_start_row=data_start_row)


def _format_sheet(
    worksheet, workbook, df, sheet_name,
    number_format, column_widths, freeze_panes, auto_filter,
    header_row=1, data_start_row=2,
):
    """格式化工作表"""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # 表头样式
    header_fill = PatternFill(patternType="solid", fgColor="4472C4")
    header_font = Font(bold=True, size=11, color="FFFFFF")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col_idx in range(1, len(df.columns) + 1):
        cell = worksheet.cell(row=header_row, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # 数据行样式
    data_font = Font(size=10)
    data_align = Alignment(vertical="center")
    number_fmt = number_format

    numeric_cols = []
    for col_idx, col_name in enumerate(df.columns, 1):
        if df[col_name].dtype in (np.float64, np.int64, float, int):
            numeric_cols.append(col_idx)

    for row_idx in range(data_start_row, data_start_row + len(df)):
        for col_idx in range(1, len(df.columns) + 1):
            cell = worksheet.cell(row=row_idx, column=col_idx)
            cell.font = data_font
            cell.alignment = data_align
            cell.border = thin_border

            if col_idx in numeric_cols and cell.value is not None:
                cell.number_format = number_fmt

    # 自动列宽
    if column_widths:
        for col_name, width in column_widths.items():
            if col_name in df.columns:
                col_idx = df.columns.get_loc(col_name) + 1
                worksheet.column_dimensions[get_column_letter(col_idx)].width = width
    else:
        # 自动计算列宽
        for col_idx, col_name in enumerate(df.columns, 1):
            # 计算最大宽度
            col_letter = get_column_letter(col_idx)
            max_length = len(str(col_name)) * 2  # 中文字符按2倍宽度

            for row_idx in range(data_start_row, min(data_start_row + 50, data_start_row + len(df))):
                cell_value = worksheet.cell(row=row_idx, column=col_idx).value
                if cell_value is not None:
                    cell_len = len(str(cell_value))
                    # 中文字符按2倍算
                    cjk_count = sum(1 for c in str(cell_value) if '\u4e00' <= c <= '\u9fff')
                    cell_len = cell_len + cjk_count
                    max_length = max(max_length, cell_len)

            # 设置列宽（限制最大宽度）
            worksheet.column_dimensions[col_letter].width = min(max_length + 4, 50)

    # 冻结窗格
    if freeze_panes:
        freeze_cell = f"A{header_row + 1}"
        worksheet.freeze_panes = freeze_cell

    # 自动筛选
    if auto_filter and len(df.columns) > 0:
        last_col = get_column_letter(len(df.columns))
        last_row = header_row + len(df)
        worksheet.auto_filter.ref = f"A{header_row}:{last_col}{last_row}"


# ============================================================================
# 高级导出功能
# ============================================================================

def export_multi_sheet(
    sheets_data: Dict[str, pd.DataFrame],
    file_path: Optional[Union[str, Path]] = None,
    global_title: Optional[str] = None,
) -> str:
    """
    导出多 sheet Excel 文件

    Args:
        sheets_data: {sheet名: DataFrame} 字典
        file_path: 输出文件路径
        global_title: 全局标题

    Returns:
        输出文件路径
    """
    if file_path is None:
        DEFAULT_OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        file_path = DEFAULT_OUTPUT_DIR / f"多表导出_{timestamp}.xlsx"

    file_path = Path(file_path)
    file_path.parent.mkdir(parents=True, exist_ok=True)

    with pd.ExcelWriter(file_path, engine="openpyxl") as writer:
        for sheet_name, df in sheets_data.items():
            # 截断 sheet 名称（Excel 限制31字符）
            safe_name = sheet_name[:31]
            export_to_excel(
                df, file_path, sheet_name=safe_name,
                title=global_title,
                freeze_panes=True, auto_filter=True,
            )

    return str(file_path)


def export_balance_sheet(
    df: pd.DataFrame,
    company_name: str,
    period: str,
    file_path: Optional[Union[str, Path]] = None,
) -> str:
    """
    导出格式化的资产负债表

    Args:
        df: 资产负债表数据
        company_name: 公司名称
        period: 期间 YYYYMM
        file_path: 输出路径

    Returns:
        输出文件路径
    """
    title = "资产负债表"
    subtitle = f"{company_name}  {_format_period_text(period)}  单位：元"

    export_df = _normalize_export_columns(
        df,
        ["资产", "行次", "期末余额", "年初余额", "负债和所有者权益", "行次2", "期末余额2", "年初余额2",
         "行次", "项目", "期末余额", "年初余额"],
    )

    col_widths = {
        "资产": 28,
        "负债和所有者权益": 28,
        "行次": 8,
        "行次2": 8,
        "项目": 30,
        "期末余额": 18,
        "年初余额": 18,
        "期末余额2": 18,
        "年初余额2": 18,
    }

    if file_path is None:
        file_path = _default_report_path("资产负债表")

    path = export_to_excel(
        export_df, file_path,
        sheet_name="资产负债表",
        title=title,
        subtitle=subtitle,
        column_widths=col_widths,
        auto_filter=False,
    )
    _apply_statement_style(path, "资产负债表")
    return path


def export_income_statement(
    df: pd.DataFrame,
    company_name: str,
    period: str,
    file_path: Optional[Union[str, Path]] = None,
) -> str:
    """
    导出格式化的利润表

    Args:
        df: 利润表数据
        company_name: 公司名称
        period: 期间
        file_path: 输出路径

    Returns:
        输出文件路径
    """
    title = "利润表"
    subtitle = f"{company_name}  {_format_period_text(period)}  单位：元"

    export_df = _normalize_export_columns(
        df,
        ["行次", "项目", "本期金额", "累计金额", "期末余额", "是否小计", "缩进层级"],
    )

    col_widths = {
        "行次": 8,
        "项目": 30,
        "本期金额": 18,
        "累计金额": 18,
        "期末余额": 18,
    }

    if file_path is None:
        file_path = _default_report_path("利润表")

    path = export_to_excel(
        export_df, file_path,
        sheet_name="利润表",
        title=title,
        subtitle=subtitle,
        column_widths=col_widths,
        auto_filter=False,
    )
    _apply_statement_style(path, "利润表")
    return path


def export_cashflow(
    df: pd.DataFrame,
    company_name: str,
    period: str,
    file_path: Optional[Union[str, Path]] = None,
) -> str:
    """导出格式化的现金流量表"""
    title = "现金流量表"
    subtitle = f"{company_name}  {_format_period_text(period)}  单位：元"

    export_df = _normalize_export_columns(
        df,
        ["行次", "项目", "期末余额", "是否小计", "缩进层级"],
    )

    col_widths = {
        "行次": 8,
        "项目": 42,
        "期末余额": 18,
        "是否小计": 10,
        "缩进层级": 10,
    }

    if file_path is None:
        file_path = _default_report_path("现金流量表")

    path = export_to_excel(
        export_df,
        file_path,
        sheet_name="现金流量表",
        title=title,
        subtitle=subtitle,
        column_widths=col_widths,
        auto_filter=False,
    )
    _apply_statement_style(path, "现金流量表")
    return path


def export_income_statement_pivot(
    pivot_df: pd.DataFrame,
    year: str,
    file_path: Optional[Union[str, Path]] = None,
) -> str:
    """
    导出损益表（透视格式，与模板一致）

    Args:
        pivot_df: 透视后的 DataFrame（第一列为"项目"，其余列为各公司数据）
        year: 年份
        file_path: 输出路径

    Returns:
        输出文件路径
    """
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl import Workbook

    if file_path is None:
        DEFAULT_OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        file_path = DEFAULT_OUTPUT_DIR / f"损益表_{ts}.xlsx"

    file_path = Path(file_path)
    file_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = f"{year}"

    # 样式
    title_font = Font(name="微软雅黑", size=14, bold=True)
    header_font = Font(name="微软雅黑", size=10, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    data_font = Font(name="微软雅黑", size=10)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    center_align = Alignment(horizontal="center", vertical="center")

    # R1: 标题
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(pivot_df.columns))
    ws.cell(1, 1, f"{year}年损益表").font = title_font
    ws.cell(1, 1).alignment = Alignment(horizontal="center", vertical="center")

    # R2: 副标题
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(pivot_df.columns))
    ws.cell(2, 1, f"单位名称：多维教育").font = Font(name="微软雅黑", size=10, color="666666")
    ws.cell(2, 1).alignment = Alignment(horizontal="center", vertical="center")

    # R3: 空行
    # R4: 表头
    header_row = 4
    for ci, col_name in enumerate(pivot_df.columns):
        cell = ws.cell(header_row, ci + 1, str(col_name))
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = center_align

    # R5+: 数据行
    for ri, (_, row) in enumerate(pivot_df.iterrows()):
        excel_row = header_row + 1 + ri
        item_name = str(row.iloc[0]) if len(row) > 0 else ""
        is_percent_row = "毛利" in item_name or "净利率" in item_name
        for ci, col_name in enumerate(pivot_df.columns):
            val = row[col_name]
            cell = ws.cell(excel_row, ci + 1)
            if ci == 0:
                cell.value = str(val)
                cell.font = Font(name="微软雅黑", size=10, bold="合计" in str(val) or "计" in str(val))
            else:
                parsed = _coerce_excel_number(val, is_percent=is_percent_row)
                cell.value = parsed if parsed is not None else 0.0
                cell.number_format = '0.00%' if is_percent_row else '#,##0.00'
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center" if ci > 0 else "left", vertical="center")

    # 列宽
    ws.column_dimensions["A"].width = 30
    for ci in range(2, len(pivot_df.columns) + 1):
        ws.column_dimensions[get_column_letter(ci)].width = 16

    # 冻结表头
    ws.freeze_panes = f"A{header_row + 1}"

    wb.save(str(file_path))
    _apply_statement_style(file_path, ws.title, header_row=header_row)
    return str(file_path)


def export_account_balance(
    df: pd.DataFrame,
    company_name: str,
    period: str,
    file_path: Optional[Union[str, Path]] = None,
) -> str:
    """
    导出格式化的科目余额表

    Args:
        df: 科目余额表数据
        company_name: 公司名称
        period: 期间
        file_path: 输出路径

    Returns:
        输出文件路径
    """
    year = period[:4]
    month = period[4:6]
    title = f"科目余额表"
    subtitle = f"{company_name}  {year}年{month}月  单位：元"

    col_widths = {
        "科目编码": 12,
        "科目名称": 25,
        "期初余额": 16,
        "借方发生额": 16,
        "贷方发生额": 16,
        "期末余额": 16,
        "方向": 8,
    }

    return export_to_excel(
        df, file_path,
        sheet_name="科目余额表",
        title=title,
        subtitle=subtitle,
        column_widths=col_widths,
    )


def export_comparison_report(
    dfs: Dict[str, pd.DataFrame],
    company_name: str,
    report_name: str,
    file_path: Optional[Union[str, Path]] = None,
) -> str:
    """
    导出对比分析报表（多期间对比）

    Args:
        dfs: {期间: DataFrame} 字典
        company_name: 公司名称
        report_name: 报表名称
        file_path: 输出路径

    Returns:
        输出文件路径
    """
    from openpyxl import Workbook
    from openpyxl.utils.dataframe import dataframe_to_rows

    if file_path is None:
        DEFAULT_OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        file_path = DEFAULT_OUTPUT_DIR / f"{report_name}_对比_{timestamp}.xlsx"

    file_path = Path(file_path)
    file_path.parent.mkdir(parents=True, exist_ok=True)

    # 合并所有期间的数据
    combined = None
    for period, df in dfs.items():
        df = df.copy()
        df["期间"] = period
        if combined is None:
            combined = df
        else:
            combined = pd.concat([combined, df], ignore_index=True)

    if combined is not None:
        return export_to_excel(
            combined, file_path,
            sheet_name=report_name,
            title=f"{company_name} - {report_name}（多期间对比）",
            subtitle="单位：元",
        )

    return str(file_path)
