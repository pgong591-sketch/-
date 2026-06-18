"""Budget workbook import helpers.

The budget workbook is a business planning source, not a calculation engine.
This module extracts the confirmed annual targets and any manually maintained
monthly actuals into normalized tables so dashboards can combine them with the
warehouse facts.
"""

from __future__ import annotations

from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional

import openpyxl

from .company_aliases import resolve_company_code
from .db_connection import get_connection


CONFIRMED_BUDGET_SHEET = "年度预算完成度 (确定)"
SUPPLEMENT_BUDGET_SHEET = "年度预算完成度"
MONTH_COLUMNS = range(1, 13)
EXCLUDED_BUDGET_PROJECTS = {"中考全日制", "哆哆托育", "西平高中"}


@dataclass
class BudgetImportResult:
    """Summary returned by the budget importer."""

    success: bool
    file_path: str
    target_rows: int = 0
    actual_rows: int = 0
    unmatched_projects: List[str] = field(default_factory=list)
    errors: List[str] = field(default_factory=list)

    def to_dict(self) -> Dict[str, Any]:
        return {
            "success": self.success,
            "file_path": self.file_path,
            "target_rows": self.target_rows,
            "actual_rows": self.actual_rows,
            "unmatched_projects": sorted(set(self.unmatched_projects)),
            "errors": self.errors,
        }


def ensure_budget_tables() -> None:
    """Create budget tables for older SQLite databases."""
    conn = get_connection()
    try:
        conn.executescript(
            """
            CREATE TABLE IF NOT EXISTS budget_targets (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                budget_year TEXT NOT NULL,
                module TEXT,
                project TEXT NOT NULL,
                company_code TEXT,
                target_type TEXT NOT NULL,
                annual_target REAL DEFAULT 0,
                source_sheet TEXT,
                source_row INTEGER,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                UNIQUE(budget_year, project, target_type)
            );

            CREATE TABLE IF NOT EXISTS budget_actual_overrides (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                budget_year TEXT NOT NULL,
                period TEXT NOT NULL,
                module TEXT,
                project TEXT NOT NULL,
                company_code TEXT,
                metric_type TEXT NOT NULL,
                amount REAL DEFAULT 0,
                source_sheet TEXT,
                source_row INTEGER,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                UNIQUE(budget_year, period, project, metric_type)
            );

            CREATE INDEX IF NOT EXISTS idx_budget_targets_year_type
                ON budget_targets(budget_year, target_type);
            CREATE INDEX IF NOT EXISTS idx_budget_targets_company
                ON budget_targets(company_code);
            CREATE INDEX IF NOT EXISTS idx_budget_actual_period_type
                ON budget_actual_overrides(period, metric_type);
            CREATE INDEX IF NOT EXISTS idx_budget_actual_company
                ON budget_actual_overrides(company_code);
            """
        )
        conn.commit()
    finally:
        conn.close()


def import_budget_workbook(
    file_path: str,
    budget_year: Optional[str] = None,
    replace: bool = True,
) -> Dict[str, Any]:
    """
    Import confirmed targets and monthly actual overrides from a budget workbook.

    Args:
        file_path: Workbook path.
        budget_year: Budget year. Defaults to the first four digits in the file
            name, then falls back to the current confirmed workbook year.
        replace: Whether to replace existing imported rows for this year.
    """
    ensure_budget_tables()
    path = Path(file_path)
    result = BudgetImportResult(success=False, file_path=str(path))
    if not path.exists():
        result.errors.append(f"预算文件不存在: {path}")
        return result.to_dict()

    budget_year = budget_year or _guess_budget_year(path)
    wb = openpyxl.load_workbook(path, data_only=True, read_only=False)
    if CONFIRMED_BUDGET_SHEET not in wb.sheetnames:
        result.errors.append(f"缺少确认版 sheet: {CONFIRMED_BUDGET_SHEET}")
        return result.to_dict()

    conn = get_connection()
    try:
        if replace:
            conn.execute("DELETE FROM budget_targets WHERE budget_year = ?", (budget_year,))
            conn.execute(
                "DELETE FROM budget_actual_overrides WHERE budget_year = ?",
                (budget_year,),
            )

        confirmed = wb[CONFIRMED_BUDGET_SHEET]
        result.target_rows += _import_income_targets(
            conn, confirmed, budget_year, CONFIRMED_BUDGET_SHEET, result
        )
        imported_targets, imported_actuals = _import_profit_block(
            conn, confirmed, budget_year, CONFIRMED_BUDGET_SHEET, result
        )
        result.target_rows += imported_targets
        result.actual_rows += imported_actuals

        if SUPPLEMENT_BUDGET_SHEET in wb.sheetnames:
            supplement = wb[SUPPLEMENT_BUDGET_SHEET]
            result.actual_rows += _import_income_actuals(
                conn, supplement, budget_year, SUPPLEMENT_BUDGET_SHEET, result
            )

        conn.commit()
        result.success = True
        return result.to_dict()
    except Exception as exc:
        conn.rollback()
        result.errors.append(str(exc))
        return result.to_dict()
    finally:
        conn.close()


def _guess_budget_year(path: Path) -> str:
    digits = "".join(ch for ch in path.name if ch.isdigit())
    if len(digits) >= 4:
        return digits[:4]
    return "2026"


def _clean_text(value: Any) -> str:
    return str(value or "").strip()


def _is_subtotal_project(project: str) -> bool:
    return _clean_text(project) in {"小计", "合计", "总计"} | EXCLUDED_BUDGET_PROJECTS


def _as_amount(value: Any) -> Optional[float]:
    if value is None or value == "":
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _resolve_project(project: str, unmatched: List[str]) -> Optional[str]:
    candidates = _project_alias_candidates(project)
    for candidate in candidates:
        code, _source = resolve_company_code(candidate)
        if code:
            return code
    unmatched.append(project)
    return None


def _project_alias_candidates(project: str) -> Iterable[str]:
    raw = _clean_text(project)
    if not raw:
        return []

    candidates = [raw]
    if raw.endswith("校区"):
        candidates.append(raw[:-2])
    else:
        candidates.append(raw + "校区")
    if raw.endswith("部"):
        candidates.append(raw[:-1])
    if "总部校区" in raw:
        candidates.append(raw.replace("总部校区", "部"))
    if raw == "宏图校区":
        candidates.append("南城宏图")
    if raw == "小学总部校区":
        candidates.append("莞城小学部")
    if raw == "初中总部校区":
        candidates.append("莞城初中部")
    if raw == "高中总部校区":
        candidates.append("莞城高中部")

    seen = set()
    ordered = []
    for item in candidates:
        item = _clean_text(item)
        if item and item not in seen:
            ordered.append(item)
            seen.add(item)
    return ordered


def _upsert_target(
    conn,
    budget_year: str,
    module: str,
    project: str,
    company_code: Optional[str],
    target_type: str,
    annual_target: float,
    source_sheet: str,
    source_row: int,
) -> int:
    conn.execute(
        """
        INSERT INTO budget_targets
            (budget_year, module, project, company_code, target_type,
             annual_target, source_sheet, source_row, updated_at)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, CURRENT_TIMESTAMP)
        ON CONFLICT(budget_year, project, target_type) DO UPDATE SET
            module = excluded.module,
            company_code = excluded.company_code,
            annual_target = excluded.annual_target,
            source_sheet = excluded.source_sheet,
            source_row = excluded.source_row,
            updated_at = CURRENT_TIMESTAMP
        """,
        (
            budget_year,
            module,
            project,
            company_code,
            target_type,
            annual_target,
            source_sheet,
            source_row,
        ),
    )
    return 1


def _upsert_actual(
    conn,
    budget_year: str,
    period: str,
    module: str,
    project: str,
    company_code: Optional[str],
    metric_type: str,
    amount: float,
    source_sheet: str,
    source_row: int,
) -> int:
    conn.execute(
        """
        INSERT INTO budget_actual_overrides
            (budget_year, period, module, project, company_code, metric_type,
             amount, source_sheet, source_row, updated_at)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, CURRENT_TIMESTAMP)
        ON CONFLICT(budget_year, period, project, metric_type) DO UPDATE SET
            module = excluded.module,
            company_code = excluded.company_code,
            amount = excluded.amount,
            source_sheet = excluded.source_sheet,
            source_row = excluded.source_row,
            updated_at = CURRENT_TIMESTAMP
        """,
        (
            budget_year,
            period,
            module,
            project,
            company_code,
            metric_type,
            amount,
            source_sheet,
            source_row,
        ),
    )
    return 1


def _import_income_targets(conn, ws, budget_year: str, sheet_name: str, result: BudgetImportResult) -> int:
    rows = 0
    module = ""
    for row in range(3, ws.max_row + 1):
        module = _clean_text(ws.cell(row, 1).value) or module
        project = _clean_text(ws.cell(row, 2).value)
        target = _as_amount(ws.cell(row, 3).value)
        if not project or _is_subtotal_project(project) or target is None:
            continue
        company_code = _resolve_project(project, result.unmatched_projects)
        rows += _upsert_target(
            conn, budget_year, module, project, company_code,
            "income", target, sheet_name, row
        )
    return rows


def _import_profit_block(conn, ws, budget_year: str, sheet_name: str, result: BudgetImportResult) -> tuple[int, int]:
    target_rows = 0
    actual_rows = 0
    module = ""
    for row in range(3, ws.max_row + 1):
        module = _clean_text(ws.cell(row, 21).value) or module
        project = _clean_text(ws.cell(row, 22).value)
        target = _as_amount(ws.cell(row, 23).value)
        if not project or _is_subtotal_project(project):
            continue
        company_code = _resolve_project(project, result.unmatched_projects)
        if target is not None:
            target_rows += _upsert_target(
                conn, budget_year, module, project, company_code,
                "profit", target, sheet_name, row
            )
        for month, col in zip(MONTH_COLUMNS, range(24, 36)):
            amount = _as_amount(ws.cell(row, col).value)
            if amount is None:
                continue
            actual_rows += _upsert_actual(
                conn, budget_year, f"{budget_year}{month:02d}", module,
                project, company_code, "profit", amount, sheet_name, row
            )
    return target_rows, actual_rows


def _import_income_actuals(conn, ws, budget_year: str, sheet_name: str, result: BudgetImportResult) -> int:
    actual_rows = 0
    module = ""
    for row in range(3, ws.max_row + 1):
        module = _clean_text(ws.cell(row, 1).value) or module
        project = _clean_text(ws.cell(row, 2).value)
        if not project or _is_subtotal_project(project):
            continue
        company_code = _resolve_project(project, result.unmatched_projects)
        for month, col in zip(MONTH_COLUMNS, range(4, 16)):
            amount = _as_amount(ws.cell(row, col).value)
            if amount is None:
                continue
            actual_rows += _upsert_actual(
                conn, budget_year, f"{budget_year}{month:02d}", module,
                project, company_code, "income", amount, sheet_name, row
            )
    return actual_rows
