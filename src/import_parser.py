"""
报表解析器模块

提供各类型 Excel 报表的解析功能。
每种报表类型对应一个解析器类，负责：
1. 读取 Excel 文件
2. 识别报表类型
3. 将 Excel 数据映射为数据库表字段
4. 返回标准化的 DataFrame
"""

import re
import os
from pathlib import Path
from typing import Optional, List, Dict, Any, Tuple
from datetime import datetime

import pandas as pd
import numpy as np

from .report_types import FILE_NAME_PATTERNS, HEADER_KEYWORDS, \
    RT_ACCOUNT_BALANCE, RT_BALANCE_SHEET, RT_INCOME_STATEMENT, RT_PL_DETAIL, \
    RT_INCOME_COST_EXPENSE, RT_REVENUE_VOLUME, RT_NON_SUBJECT_ALLOCATION, RT_MGMT_DEPT_INCOME_COST, \
    RT_NON_SUBJECT_MGMT_DEPT_INCOME_COST, RT_NON_SUBJECT_TEACHING_FEE


def _clean_text(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).replace("\u3000", " ")
    return re.sub(r"\s+", "", text).strip()


def _to_float(value: Any) -> float:
    if value is None:
        return 0.0
    if isinstance(value, (int, float, np.number)):
        if pd.isna(value):
            return 0.0
        return float(value)
    text = str(value).strip()
    if not text:
        return 0.0
    text = text.replace(",", "")
    text = text.replace("（", "(").replace("）", ")")
    if text.startswith("(") and text.endswith(")"):
        text = "-" + text[1:-1]
    try:
        return float(text)
    except ValueError:
        return 0.0


def _extract_period_yyyymm(text: str) -> str:
    if not text:
        return ""
    m = re.search(r"(20\d{2})[.\-/年]?\s*(\d{1,2})", text)
    if m:
        month = int(m.group(2))
        if 1 <= month <= 12:
            return f"{m.group(1)}{month:02d}"
    m2 = re.search(r"(20\d{2})(0[1-9]|1[0-2])", text)
    if m2:
        return m2.group(1) + m2.group(2)
    return ""


def _sanitize_company_candidate(company: str) -> str:
    text = _clean_text(company)
    if not text:
        return ""
    if text in {"非学科", "课酬", "收入人次", "管理中心"}:
        return ""
    noise_keywords = [
        "管理中心部门收入成本费用表",
        "非学科管理中心部门收入成本费用表",
        "收入成本费用表",
        "非学科课酬",
        "收入人次表",
        "课酬",
    ]
    if any(k in text for k in noise_keywords):
        return ""
    return str(company).strip()


def _looks_like_matrix_dept_report(df: Optional[pd.DataFrame]) -> bool:
    if df is None or df.empty:
        return False
    sample = df.head(12).fillna("").astype(str)
    text = "".join(sample.values.flatten().tolist())
    return ("会计科目" in text and "部门档案" in text) or ("会计科目" in text and "统计方式" in text and "余额方向" in text)


def identify_report_type(
    file_path: str,
    df: Optional[pd.DataFrame] = None,
    source_name: Optional[str] = None,
) -> Optional[str]:
    """
    根据文件名和表头内容识别报表类型

    Args:
        file_path: Excel 文件路径
        df: 已读取的 DataFrame（可选）

    Returns:
        报表类型中文名，无法识别返回 None
    """
    file_name = Path(file_path).name
    candidate_names = [n for n in [source_name, file_name] if n]
    name_hint = (source_name or file_name).lower()
    zh_name_hint = "".join(candidate_names)

    # 0. ASCII 文件名兜底（处理临时拷贝文件名）
    if "非学科管理中心" in zh_name_hint:
        return RT_NON_SUBJECT_MGMT_DEPT_INCOME_COST
    if "管理中心部门收入成本费用表" in zh_name_hint:
        return RT_MGMT_DEPT_INCOME_COST
    if "非学科课酬" in zh_name_hint:
        return RT_NON_SUBJECT_TEACHING_FEE
    if "收入人次" in zh_name_hint:
        return RT_REVENUE_VOLUME

    if "non_subject_teaching_fee" in name_hint:
        return RT_NON_SUBJECT_TEACHING_FEE
    if "non_subject_mgmt_dept_income_cost" in name_hint:
        return RT_NON_SUBJECT_MGMT_DEPT_INCOME_COST
    if "mgmt_dept_income_cost" in name_hint:
        return RT_MGMT_DEPT_INCOME_COST
    if "income_cost_expense" in name_hint:
        return RT_INCOME_COST_EXPENSE
    if "revenue_volume" in name_hint:
        return RT_REVENUE_VOLUME

    # 1. 根据文件名匹配（优先原始上传名）
    for name in candidate_names:
        for pattern, report_type in FILE_NAME_PATTERNS:
            if re.search(pattern, name):
                return report_type

    # 2. 矩阵部门表优先识别（避免被通用关键字误判）
    if df is not None and not df.empty:
        sample = df.head(12).fillna("").astype(str)
        text = "".join(sample.values.flatten().tolist())
        if "收入成本费用表" in text and "部门档案" not in text and "管理公司收入成本费用表" not in text:
            return RT_INCOME_COST_EXPENSE

    if _looks_like_matrix_dept_report(df):
        sample = df.head(12).fillna("").astype(str)
        text = "".join(sample.values.flatten().tolist())
        if "素质管理中心" in text or "素质中心教研部" in text or "学术中心" in text:
            return RT_NON_SUBJECT_MGMT_DEPT_INCOME_COST
        return RT_MGMT_DEPT_INCOME_COST

    # 3. 课酬表优先识别
    if df is not None and not df.empty:
        sample = df.head(12).fillna("").astype(str)
        text = "".join(sample.values.flatten().tolist())
        if ("年份" in text and "月份" in text and "职员代码" in text and ("总收入" in text or "课酬" in text)):
            return RT_NON_SUBJECT_TEACHING_FEE

    # 4. 根据表头关键字匹配
    if df is not None:
        # 获取所有列名和前几行数据
        all_text = " ".join([str(c) for c in df.columns])
        for _, row in df.head(10).iterrows():
            all_text += " " + " ".join([str(v) for v in row.values])

        for report_type, keywords in HEADER_KEYWORDS.items():
            match_count = sum(1 for kw in keywords if kw in all_text)
            if match_count >= len(keywords) * 0.6:  # 60% 以上关键字匹配
                return report_type

    return None


# ============================================================================
# 解析器基类
# ============================================================================

class BaseParser:
    """报表解析器基类"""

    # 目标数据库表名
    TABLE_NAME: str = ""

    # 必填字段
    REQUIRED_COLUMNS: List[str] = []

    def __init__(self, company_code: Optional[str] = None, period: Optional[str] = None):
        self.company_code = company_code
        self.period = period
        self.errors: List[str] = []
        self.warnings: List[str] = []

    def parse(self, file_path: str, **kwargs) -> pd.DataFrame:
        """
        解析 Excel 文件，返回标准化 DataFrame

        Args:
            file_path: Excel 文件路径
            **kwargs: 额外参数（如指定 company_code, period）

        Returns:
            标准化后的 DataFrame
        """
        raise NotImplementedError("子类必须实现 parse 方法")

    def _read_excel(self, file_path: str, **kwargs) -> pd.DataFrame:
        """读取 Excel 文件"""
        return pd.read_excel(file_path, **kwargs)

    def _clean_df(self, df: pd.DataFrame) -> pd.DataFrame:
        """清洗 DataFrame：去除空行空列、去除前后空格"""
        # 去除全空行
        df = df.dropna(how="all")
        # 去除全空列
        df = df.dropna(axis=1, how="all")
        # 去除列名和字符串值的前后空格
        df.columns = [str(c).strip() if isinstance(c, str) else c for c in df.columns]
        for col in df.select_dtypes(include=["object", "string"]).columns:
            df[col] = df[col].map(lambda v: v.strip() if isinstance(v, str) else v)
        return df

    def _normalize_numeric(self, series: pd.Series) -> pd.Series:
        """
        将数字列标准化：去除千分位、货币符号等，转为 float

        Args:
            series: 原始数据列

        Returns:
            标准化后的数值列
        """
        if series.dtype in (np.float64, np.int64):
            return series.astype(float)

        # 字符串清洗
        result = series.astype(str)
        result = result.str.replace(",", "", regex=False)
        result = result.str.replace("￥", "", regex=False)
        result = result.str.replace("¥", "", regex=False)
        result = result.str.replace(" ", "", regex=False)
        result = result.str.replace("--", "0", regex=False)
        result = result.str.replace("（", "-", regex=False)
        result = result.str.replace("）", "", regex=False)
        result = result.str.replace("(", "-", regex=False)
        result = result.str.replace(")", "", regex=False)
        result = pd.to_numeric(result, errors="coerce").fillna(0.0)
        return result

    def _extract_company_period(self, file_path: str, df: pd.DataFrame) -> Tuple[Optional[str], Optional[str]]:
        """
        从文件路径或数据中提取公司和期间

        优先级: 构造函数传入 > 文件路径 > 数据内容

        Args:
            file_path: 文件路径
            df: 数据 DataFrame

        Returns:
            (company_code, period)
        """
        company = self.company_code
        period = self.period

        # 尝试从文件路径提取
        if not company or not period:
            file_name = Path(file_path).name
            # 模式1: "公司名_202603_科目余额表" (文字开头)
            m = re.search(r"(.+?)[_\- ](\d{6})", file_name)
            if m:
                if not company:
                    company = m.group(1).strip()
                if not period:
                    period = m.group(2).strip()
            else:
                # 模式2: "202603公司名_科目余额表.xlsx" (数字开头)
                m = re.search(r"^(\d{6})[\s_\-]*(.+?)(?:\.xlsx?|$)", file_name)
                if m:
                    if not period:
                        period = m.group(1).strip()
                    if not company:
                        raw = m.group(2).strip()
                        # 去除报表类型关键词，提取公司名
                        for kw in ["科目辅助余额表", "科目辅助余额", "科目余额表",
                                    "辅助余额表", "辅助余额", "余额表",
                                    "资产负债表", "利润表", "损益表",
                                    "收入成本费用表", "收入成本费用明细表",
                                    "现金流量表", "收入人次表", "收入人次", 
                                    "课酬表", "课酬", ".xlsx", ".xls"]:
                            raw = raw.replace(kw, "")
                        # 提取连续中文字符作为公司名
                        name_match = re.search(r"[\u4e00-\u9fff]+", raw)
                        if name_match:
                            company = name_match.group()

        return company, period

    def get_import_result(self) -> Dict[str, Any]:
        """获取导入结果信息"""
        return {
            "errors": self.errors,
            "warnings": self.warnings,
            "is_valid": len(self.errors) == 0,
        }


# ============================================================================
# 科目余额表解析器
# ============================================================================

class AccountBalanceParser(BaseParser):
    """科目余额表解析器"""

    TABLE_NAME = "account_balance"
    REQUIRED_COLUMNS = ["account_code", "account_name", "opening_balance",
                         "debit_amount", "credit_amount", "ending_balance"]

    # 常见 Excel 表头映射（含辅助余额表变体）
    COLUMN_MAPPING = {
        # 科目编码
        "科目编码": "account_code",
        "科目代码": "account_code",
        "科目编号": "account_code",
        "会计科目编码": "account_code",
        "会计科目代码": "account_code",
        # 科目名称
        "科目名称": "account_name",
        "会计科目名称": "account_name",
        "会计科目": "account_name",
        # 期初/年初
        "期初余额": "opening_balance",
        "年初余额": "opening_balance",
        "期初": "opening_balance",
        "期初借方余额": "opening_balance",
        "年初借方余额": "opening_balance",
        # 本期借方
        "本期借方": "debit_amount",
        "本期借方发生额": "debit_amount",
        "借方": "debit_amount",
        "借方发生额": "debit_amount",
        "借方金额": "debit_amount",
        "本月借方": "debit_amount",
        # 本期贷方
        "本期贷方": "credit_amount",
        "本期贷方发生额": "credit_amount",
        "贷方": "credit_amount",
        "贷方发生额": "credit_amount",
        "贷方金额": "credit_amount",
        "本月贷方": "credit_amount",
        # 期末
        "期末余额": "ending_balance",
        "期末": "ending_balance",
        "期末借方余额": "ending_balance",
        "期末贷方余额": "ending_balance",
        "期末借方": "ending_balance",
        "期末贷方": "ending_balance",
        # 方向
        "余额方向": "direction",
        "方向": "direction",
        "借或贷": "direction",
        # 辅助核算
        "辅助核算": "assist_dimensions",
        "辅助项目": "assist_dimensions",
        "项目": "assist_dimensions",
    }

    def parse(self, file_path: str, **kwargs) -> pd.DataFrame:
        """
        解析科目余额表 Excel

        支持常见格式：
        - 标准格式：科目编码 | 科目名称 | 期初余额 | 借方 | 贷方 | 期末余额
        - 带表头的复杂格式（如辅助余额表）
        - "科目编码\科目名称" 合并在同一列的格式

        Args:
            file_path: Excel 文件路径
            **kwargs: 额外参数

        Returns:
            标准化的 DataFrame
        """
        df_raw = self._read_excel(file_path, header=None)

        company, period = self._extract_company_period(file_path, df_raw)
        # 注意：kwargs.get("company_code") 在值为 None 时会返回 None，不能覆盖已提取的值
        if kwargs.get("company_code") is not None:
            company = kwargs["company_code"]
        if kwargs.get("period") is not None:
            period = kwargs["period"]

        # === 特殊处理：尝试从标题行提取期间（如 "期间：2026.03-2026.03"）===
        if not period:
            for i in range(min(5, len(df_raw))):
                row_text = " ".join([str(v) for v in df_raw.iloc[i].values if pd.notna(v)])
                m = re.search(r"期间[：:]\s*(\d{4})[.\-](\d{1,2})", row_text)
                if m:
                    period = m.group(1) + m.group(2).zfill(2)
                    break

        # === 特殊处理：用 "科目辅助余额表" 的表头（第4行 = 索引3）===
        header_row = self._find_header_row(df_raw)

        if header_row is not None:
            df = self._read_excel(file_path, header=header_row)
        else:
            best_row = self._try_find_header_by_mapping(df_raw)
            if best_row is not None:
                df = self._read_excel(file_path, header=best_row)
            else:
                df = df_raw

        df = self._clean_df(df)

        # === 特殊处理：去重并列的列名（如"方向"出现两次）===
        cols = list(df.columns)
        seen = {}
        new_cols = []
        for c in cols:
            c_str = str(c).strip()
            if c_str in seen:
                seen[c_str] += 1
                new_cols.append(f"{c_str}_{seen[c_str]}")
            else:
                seen[c_str] = 0
                new_cols.append(c_str)
        df.columns = new_cols

        # 映射列名
        df = self._map_columns(df)

        # === 特殊处理：科目编码和名称合并在同一列（如 "1001\现金"）===
        # 检查 account_name 列是否包含编码前缀
        backslash = chr(92)  # 单反斜杠
        if "account_name" in df.columns:
            sample_vals = df["account_name"].dropna().astype(str).head(5).tolist()
            has_combined = any(backslash in v for v in sample_vals)
            has_subtotal = any("科目合计" in v for v in sample_vals)

            if has_combined or has_subtotal:
                def clean_account_name(val):
                    if pd.isna(val):
                        return val
                    s = str(val).strip()
                    # 去掉 Excel 文本格式前导单引号
                    if s.startswith("'") and len(s) > 1:
                        s = s[1:]
                    if backslash in s:
                        parts = s.split(backslash, 1)
                        return parts[1].strip() if len(parts) > 1 else s
                    return s

                # 如果 account_code 为空，从 account_name 中提取
                if "account_code" in df.columns:
                    code_from_first_col = df["account_code"].dropna().astype(str).str.strip()
                    code_from_first_col = code_from_first_col[code_from_first_col != ""]
                    if len(code_from_first_col) == 0:
                        # 从 account_name 中提取编码
                        df["account_code"] = df["account_name"].astype(str).apply(
                            lambda v: v.split(backslash)[0].strip().strip("'") if backslash in str(v) else v
                        )

                # 清洗 account_name（去掉编码前缀）
                df["account_name"] = df["account_name"].apply(clean_account_name)

        if "account_code" not in df.columns or df["account_code"].isna().all():
            df_guess = self._guess_columns_by_position(df)
            if df_guess is not None:
                df = df_guess

        # 标准化数值
        for col in ["opening_balance", "debit_amount", "credit_amount", "ending_balance"]:
            if col in df.columns:
                df[col] = self._normalize_numeric(df[col])

        # 填充公司、期间
        df["company_code"] = company or ""
        df["period"] = period or ""

        # 填充余额方向
        if "direction" not in df.columns:
            df["direction"] = "借"

        # 如果有辅助核算列，存入 assist_dimensions
        if "assist_dimensions" in df.columns:
            def fmt_assist(v):
                if pd.isna(v) or str(v).strip() in ("", "0"):
                    return ""
                return str(v).strip()
            df["assist_dimensions"] = df["assist_dimensions"].apply(fmt_assist)

        # 计算期末余额（如果没有）
        if "ending_balance" not in df.columns or df["ending_balance"].isna().all():
            if all(c in df.columns for c in ["opening_balance", "debit_amount", "credit_amount"]):
                df["ending_balance"] = df["opening_balance"] + df["debit_amount"] - df["credit_amount"]

        # 选择标准列
        result_cols = ["company_code", "period", "account_code", "account_name",
                        "opening_balance", "debit_amount", "credit_amount",
                        "ending_balance", "direction", "assist_dimensions"]
        available_cols = [c for c in result_cols if c in df.columns]
        df_result = df[available_cols].copy()

        # 过滤掉汇总行、制表人、核算单位等非数据行
        if "account_code" in df_result.columns:
            skip_patterns = [
                "科目合计", "合计", "总计", "小计",  # 汇总行
                "制表人", "核算单位",                # 页脚信息
                "制单人",                           # 制单人
                "打印时间",                         # 打印时间
                "科目编码",                         # 表头行混入数据
            ]
            mask = pd.Series(True, index=df_result.index)
            for pat in skip_patterns:
                mask &= ~df_result["account_code"].astype(str).str.contains(pat, na=False, regex=False)
            # 同时过滤 account_name 中的非数据行
            if "account_name" in df_result.columns:
                for pat in skip_patterns:
                    mask &= ~df_result["account_name"].astype(str).str.contains(pat, na=False, regex=False)
            # 过滤掉关键字段全空的行
            key_cols = [c for c in ["account_code", "account_name"] if c in df_result.columns]
            if key_cols:
                mask &= df_result[key_cols].notna().any(axis=1)
            df_result = df_result[mask]

        return df_result

    def _find_header_row(self, df: pd.DataFrame) -> Optional[int]:
        """查找表头行位置"""
        keywords = ["科目编码", "科目名称", "科目代码", "期初余额", "期初",
                     "借方", "贷方", "期末余额", "会计科目", "科目",
                     "年初余额", "期末余额"]

        for idx, row in df.iterrows():
            row_text = " ".join([str(v) for v in row.values])
            match_count = sum(1 for kw in keywords if kw in row_text)
            if match_count >= 2:  # 至少匹配2个关键字
                return idx

        return None

    def _try_find_header_by_mapping(self, df: pd.DataFrame) -> Optional[int]:
        """通过列名映射试探找出真正的表头行"""
        for try_row in range(min(10, len(df))):
            try:
                row = df.iloc[try_row]
                # 检查这一行映射后能识别出几个标准字段
                mapped = 0
                for cell in row:
                    cell_str = str(cell).strip()
                    if cell_str in self.COLUMN_MAPPING:
                        target = self.COLUMN_MAPPING[cell_str]
                        if target in ("account_code", "account_name",
                                      "opening_balance", "ending_balance",
                                      "debit_amount", "credit_amount"):
                            mapped += 1
                if mapped >= 3:  # 至少匹配3个关键字段
                    return try_row
            except Exception:
                continue
        return None

    def _guess_columns_by_position(self, df: pd.DataFrame) -> Optional[pd.DataFrame]:
        """按常见列位置猜测字段（适应"科目辅助余额表"等变体）"""
        if len(df.columns) < 4:
            return None

        result = df.copy()
        ncols = len(result.columns)
        col_map = {}

        # 常见科目余额表的列位置顺序：
        # 科目编码 | 科目名称 | 期初余额 | 借方 | 贷方 | 期末余额
        # 科目编码 | 科目名称 | 辅助核算 | 期初 | 借方 | 贷方 | 期末
        col_map[result.columns[0]] = "account_code"
        if ncols >= 2:
            col_map[result.columns[1]] = "account_name"

        # 从第3列开始判断哪列是期初/借方/贷方/期末
        remaining = list(result.columns[2:])
        if len(remaining) >= 4:
            col_map[remaining[0]] = "opening_balance"
            col_map[remaining[1]] = "debit_amount"
            col_map[remaining[2]] = "credit_amount"
            col_map[remaining[3]] = "ending_balance"
        elif len(remaining) == 3:
            col_map[remaining[0]] = "opening_balance"
            col_map[remaining[1]] = "debit_amount"
            col_map[remaining[2]] = "credit_amount"

        result = result.rename(columns=col_map)

        # 检查是否成功
        if "account_code" in result.columns and "account_name" in result.columns:
            return result
        return None

    def _map_columns(self, df: pd.DataFrame) -> pd.DataFrame:
        """将中文列名映射为标准字段名"""
        col_mapping = {}
        for col in df.columns:
            col_str = str(col).strip()
            if col_str in self.COLUMN_MAPPING:
                col_mapping[col_str] = self.COLUMN_MAPPING[col_str]
        return df.rename(columns=col_mapping)


# ============================================================================
# 损益明细表解析器
# ============================================================================

class PlDetailParser(BaseParser):
    """损益明细表（收入成本费用明细表）解析器"""

    TABLE_NAME = "pl_detail"
    REQUIRED_COLUMNS = ["item_code", "item_name", "category", "amount"]

    COLUMN_MAPPING = {
        "项目编码": "item_code",
        "项目代码": "item_code",
        "项目编号": "item_code",
        "项目名称": "item_name",
        "收入项目": "item_name",
        "成本项目": "item_name",
        "费用项目": "item_name",
        "类别": "category",
        "项目类别": "category",
        "金额": "amount",
        "本期金额": "amount",
        "本月金额": "amount",
        "部门编码": "dept_code",
        "部门代码": "dept_code",
        "部门": "dept_name",
        "部门名称": "dept_name",
        "备注": "remark",
    }

    SUMMARY_NAMES = {"收入合计", "成本费用合计", "净利润", "折旧费", "待摊费", "折旧与待摊费用合计", "净利润（不含折旧与摊销）"}

    def parse(self, file_path: str, **kwargs) -> pd.DataFrame:
        """解析损益明细表"""
        special = self._parse_income_cost_expense_workbook(file_path, **kwargs)
        if special is not None:
            return special

        df = self._read_excel(file_path)
        df = self._clean_df(df)
        df = df.rename(columns=self.COLUMN_MAPPING)

        company = kwargs.get("company_code") or self.company_code or ""
        period = kwargs.get("period") or self.period or ""

        df["company_code"] = company
        df["period"] = period

        if "amount" in df.columns:
            df["amount"] = self._normalize_numeric(df["amount"])

        # 如果缺少 item_code，使用 item_name 代替
        if "item_code" not in df.columns and "item_name" in df.columns:
            df["item_code"] = df["item_name"]
        if "dept_code" in df.columns:
            df["dept_code"] = df["dept_code"].fillna("").astype(str).str.strip()

        result_cols = ["company_code", "period", "item_code", "item_name",
                        "category", "amount", "dept_code", "dept_name", "remark"]
        available_cols = [c for c in result_cols if c in df.columns]
        return df[available_cols].copy()

    def _parse_income_cost_expense_workbook(self, file_path: str, **kwargs) -> Optional[pd.DataFrame]:
        """Parse the formatted income/cost/expense workbook used by operating summary."""
        raw = pd.read_excel(file_path, sheet_name=0, header=None)
        if raw.empty:
            return None
        sample_text = "".join(raw.head(12).fillna("").astype(str).values.flatten().tolist())
        if "收入成本费用表" not in sample_text:
            return None

        header_idx = self._find_income_cost_header(raw)
        if header_idx is None:
            self.errors.append("未找到收入成本费用表表头行")
            return pd.DataFrame()

        header = raw.iloc[header_idx].fillna("").astype(str).str.strip().tolist()
        company = self._extract_income_cost_company(raw, kwargs.get("company_code") or self.company_code)
        period = kwargs.get("period") or self.period or self._extract_income_cost_period(raw, file_path)
        month = int(str(period)[4:6]) if str(period).isdigit() and len(str(period)) == 6 else None
        current_col = self._find_header_col(header, f"{month}月") if month else None
        ytd_col = self._find_header_col(header, "本年累计")
        code_col = self._find_header_col(header, "科目代码")
        name_col = self._find_header_col(header, "科目")
        if current_col is None or ytd_col is None or name_col is None:
            self.errors.append("收入成本费用表缺少期间列、科目列或本年累计列")
            return pd.DataFrame()

        rows = []
        income_section = True
        for row_idx in range(header_idx + 1, len(raw)):
            row = raw.iloc[row_idx]
            name = str(row.iloc[name_col]).strip() if name_col < len(row) and not pd.isna(row.iloc[name_col]) else ""
            code = str(row.iloc[code_col]).strip() if code_col is not None and code_col < len(row) and not pd.isna(row.iloc[code_col]) else ""
            current_value = _to_float(row.iloc[current_col]) if current_col < len(row) else 0.0
            ytd_value = _to_float(row.iloc[ytd_col]) if ytd_col < len(row) else 0.0
            if not name and not code:
                if rows:
                    break
                continue
            if name == "收入合计":
                income_section = False
            if abs(current_value) < 1e-9 and abs(ytd_value) < 1e-9 and name not in self.SUMMARY_NAMES:
                continue
            item_code = code or f"SUMMARY_{len(rows) + 1:03d}_{name}"
            rows.append(
                {
                    "company_code": company,
                    "period": period,
                    "item_code": item_code,
                    "item_name": name,
                    "category": self._income_cost_category(name, income_section),
                    "amount": current_value,
                    "dept_code": "",
                    "dept_name": "",
                    "remark": f"本年累计={ytd_value:.2f}; 来源=收入成本费用表第{row_idx + 1}行",
                }
            )

        if not rows:
            self.errors.append("未解析到收入成本费用表明细或汇总数据")
            return pd.DataFrame()
        return pd.DataFrame(rows)

    def _find_income_cost_header(self, raw: pd.DataFrame) -> Optional[int]:
        for idx in range(min(len(raw), 20)):
            values = [str(v).strip() for v in raw.iloc[idx].fillna("").tolist()]
            text = "".join(values)
            if "科目代码" in text and "科目" in text and "本年累计" in text:
                return idx
        return None

    def _find_header_col(self, header: list[str], name: str) -> Optional[int]:
        for idx, value in enumerate(header):
            if value == name:
                return idx
        return None

    def _extract_income_cost_period(self, raw: pd.DataFrame, file_path: str) -> str:
        text = " ".join(raw.head(8).fillna("").astype(str).values.flatten().tolist())
        period = _extract_period_yyyymm(text)
        return period or _extract_period_yyyymm(Path(file_path).name)

    def _extract_income_cost_company(self, raw: pd.DataFrame, fallback: Optional[str]) -> str:
        if fallback:
            return str(fallback).strip()
        top = raw.head(8).fillna("").astype(str)
        for _, row in top.iterrows():
            for value in row.tolist():
                text = str(value).strip()
                if re.fullmatch(r"\d{3,}", text):
                    return text
        text = " ".join(top.values.flatten().tolist())
        m = re.search(r"单位[:：]\s*([^\s]+)", text)
        return m.group(1).strip() if m else ""

    def _income_cost_category(self, item_name: str, income_section: bool) -> str:
        if item_name in {"收入合计", "主营业务收入", "其他业务收入", "投资收益", "营业外收入"} or income_section:
            return "收入"
        if "利润" in item_name:
            return "利润"
        return "费用"


# ============================================================================
# 收入人次表解析器
# ============================================================================

class RevenueVolumeParser(BaseParser):
    """收入人次表解析器"""

    TABLE_NAME = "revenue_volume"
    REQUIRED_COLUMNS = ["product_line", "customer_count", "revenue_amount"]

    COLUMN_MAPPING = {
        "产品线": "product_line",
        "产品": "product_line",
        "项目": "product_line",
        "业务类型": "product_line",
        "人次": "customer_count",
        "人数": "customer_count",
        "客户数量": "customer_count",
        "收入": "revenue_amount",
        "收入金额": "revenue_amount",
        "单价": "unit_price",
    }

    def parse(self, file_path: str, **kwargs) -> pd.DataFrame:
        """解析收入人次表"""
        company = _sanitize_company_candidate(kwargs.get("company_code") or self.company_code or "")
        period = kwargs.get("period") or self.period or ""

        # 优先解析“收入人次表”明细页（通常是第三个工作表）。
        try:
            xls = pd.ExcelFile(file_path)
            target_sheet = None
            for s in xls.sheet_names:
                s_name = str(s)
                if "收入人次表" in s_name and "年级表" not in s_name and "科目表" not in s_name:
                    target_sheet = s
                    break
            if target_sheet is None and len(xls.sheet_names) >= 3:
                target_sheet = xls.sheet_names[2]

            if target_sheet is not None:
                detail = pd.read_excel(file_path, sheet_name=target_sheet)
                detail = self._clean_df(detail)
                required_cols = {"校区名称", "年级", "科目", "人次_次数", "实际收入"}
                if required_cols.issubset(set(detail.columns)):
                    if not period:
                        period = _extract_period_yyyymm(Path(file_path).name)
                        if not period and "年份" in detail.columns and "月份" in detail.columns and len(detail) > 0:
                            year_non_null = pd.to_numeric(detail["年份"], errors="coerce").dropna()
                            month_non_null = pd.to_numeric(detail["月份"], errors="coerce").dropna()
                            if not year_non_null.empty and not month_non_null.empty:
                                period = f"{int(year_non_null.iloc[0])}{int(month_non_null.iloc[0]):02d}"

                    if not company:
                        company = "东莞非学科管理中心"
                        self.warnings.append("收入人次表未识别到公司编码，已默认使用“东莞非学科管理中心”，可在导入设置手工覆盖。")

                    campus_alias = {
                        "初中总部校区": "莞城初中部",
                        "小学总部校区": "莞城小学部",
                        "深圳福田校区": "深圳卓越",
                        "茶山校区": "茶山学前",
                        "西平小初校区": "西平校区",
                    }

                    work = detail.copy()
                    work["校区名称"] = work["校区名称"].astype(str).str.strip().replace(campus_alias)
                    work = work[work["校区名称"] != "班德校区"].copy()

                    work["customer_count"] = pd.to_numeric(work["人次_次数"], errors="coerce").fillna(0)
                    work["revenue_amount"] = pd.to_numeric(work["实际收入"], errors="coerce").fillna(0)
                    work["year"] = pd.to_numeric(work.get("年份"), errors="coerce").fillna(0).astype(int)
                    work["month"] = pd.to_numeric(work.get("月份"), errors="coerce").fillna(0).astype(int)
                    work["source_quarter_label"] = work.get("季度", "").astype(str).str.strip()
                    work["campus_name"] = work["校区名称"].astype(str).str.strip()
                    work["grade"] = work["年级"].astype(str).str.strip()
                    work["subject"] = work["科目"].astype(str).str.strip()
                    work["data_period"] = period
                    work["business_period"] = np.where(
                        work["year"] > 0,
                        work["year"].astype(str) + work["month"].clip(lower=1).astype(str).str.zfill(2),
                        "",
                    )
                    work["calendar_quarter"] = np.where(
                        work["year"] > 0,
                        work["year"].astype(str)
                        + "Q"
                        + (((work["month"].clip(lower=1) - 1) // 3) + 1).astype(str),
                        "",
                    )
                    work["source_file"] = Path(file_path).name
                    work["source_sheet"] = str(target_sheet)

                    # product_line 保留兼容用途，同时继续承载唯一性。
                    work["product_line"] = (
                        work["campus_name"]
                        + "|"
                        + work["grade"]
                        + "|"
                        + work["subject"]
                        + "|"
                        + work["business_period"].astype(str)
                        + "|"
                        + work["source_quarter_label"]
                    )

                    group_cols = [
                        "company_code",
                        "period",
                        "data_period",
                        "business_period",
                        "year",
                        "month",
                        "calendar_quarter",
                        "source_quarter_label",
                        "campus_name",
                        "grade",
                        "subject",
                        "product_line",
                        "source_file",
                        "source_sheet",
                    ]
                    work["company_code"] = company
                    work["period"] = period
                    grouped = (
                        work.groupby(group_cols, dropna=False)[["customer_count", "revenue_amount"]]
                        .sum()
                        .reset_index()
                    )
                    grouped["unit_price"] = np.where(
                        grouped["customer_count"] > 0,
                        grouped["revenue_amount"] / grouped["customer_count"],
                        0,
                    )
                    return grouped[[
                        "company_code",
                        "period",
                        "product_line",
                        "data_period",
                        "business_period",
                        "year",
                        "month",
                        "calendar_quarter",
                        "source_quarter_label",
                        "campus_name",
                        "grade",
                        "subject",
                        "customer_count",
                        "revenue_amount",
                        "unit_price",
                        "source_file",
                        "source_sheet",
                    ]]
        except Exception:
            # 回退到旧解析路径
            pass

        # 旧版平铺表头解析（兜底）
        df = self._read_excel(file_path)
        df = self._clean_df(df)
        df = df.rename(columns=self.COLUMN_MAPPING)

        df["company_code"] = company
        df["period"] = period

        for col in ["customer_count", "revenue_amount", "unit_price"]:
            if col in df.columns:
                df[col] = self._normalize_numeric(df[col])

        if "unit_price" not in df.columns and "revenue_amount" in df.columns and "customer_count" in df.columns:
            df["unit_price"] = np.where(
                df["customer_count"] > 0,
                df["revenue_amount"] / df["customer_count"],
                0
            )

        result_cols = ["company_code", "period", "product_line",
                        "data_period", "business_period", "year", "month",
                        "calendar_quarter", "source_quarter_label",
                        "campus_name", "grade", "subject",
                        "customer_count", "revenue_amount", "unit_price",
                        "source_file", "source_sheet"]
        available_cols = [c for c in result_cols if c in df.columns]
        return df[available_cols].copy()


# ============================================================================
# 非学科费用分配表解析器
# ============================================================================

class NonSubjectAllocationParser(BaseParser):
    """非学科费用分配表解析器"""

    TABLE_NAME = "non_subject_allocation"

    COLUMN_MAPPING = {
        "成本中心": "cost_center",
        "成本中心编码": "cost_center",
        "科目": "account_name",
        "科目编码": "account_code",
        "科目名称": "account_name",
        "分配基数": "allocation_base",
        "分配金额": "allocated_amount",
        "比例": "ratio",
        "分配比例": "ratio",
    }

    def parse(self, file_path: str, **kwargs) -> pd.DataFrame:
        """解析非学科费用分配表"""
        df = self._read_excel(file_path)
        df = self._clean_df(df)
        df = df.rename(columns=self.COLUMN_MAPPING)

        company = kwargs.get("company_code") or self.company_code or ""
        period = kwargs.get("period") or self.period or ""

        df["company_code"] = company
        df["period"] = period

        for col in ["allocation_base", "allocated_amount", "ratio"]:
            if col in df.columns:
                df[col] = self._normalize_numeric(df[col])
        if "account_code" in df.columns:
            df["account_code"] = df["account_code"].fillna("").astype(str).str.strip()

        result_cols = ["company_code", "period", "cost_center",
                        "account_code", "account_name",
                        "allocation_base", "allocated_amount", "ratio"]
        available_cols = [c for c in result_cols if c in df.columns]
        return df[available_cols].copy()


# ============================================================================
# 管理中心部门收入成本费用表解析器
# ============================================================================

class MgmtDeptIncomeCostParser(BaseParser):
    """管理中心部门收入成本费用表解析器"""

    TABLE_NAME = "mgmt_dept_income_cost"

    COLUMN_MAPPING = {
        "部门编码": "dept_code",
        "部门代码": "dept_code",
        "部门名称": "dept_name",
        "部门": "dept_name",
        "收入": "income_amount",
        "收入金额": "income_amount",
        "成本": "cost_amount",
        "成本金额": "cost_amount",
        "费用": "expense_amount",
        "费用金额": "expense_amount",
        "利润": "profit_amount",
        "利润金额": "profit_amount",
    }

    @staticmethod
    def _extract_dept_name(header_value: Any) -> str:
        text = _clean_text(header_value)
        if not text:
            return ""
        if "部门档案" in text:
            # 例如：【部门档案：董事会】
            text = text.replace("【", "").replace("】", "")
            text = text.replace("部门档案:", "").replace("部门档案：", "")
            return _clean_text(text)
        return text

    def _parse_matrix_layout(self, raw: pd.DataFrame, company: str, period: str) -> Optional[pd.DataFrame]:
        if not _looks_like_matrix_dept_report(raw):
            return None

        header_row = None
        for ridx in range(min(12, len(raw))):
            row_values = [_clean_text(v) for v in raw.iloc[ridx].tolist()]
            row_text = "".join(row_values)
            if "会计科目" in row_text and any("部门档案" in x for x in row_values):
                header_row = ridx
                break
        if header_row is None:
            header_row = 5 if len(raw) > 7 else 0

        dept_cols: List[Tuple[int, str]] = []
        for c in range(3, raw.shape[1]):
            dept_name = self._extract_dept_name(raw.iat[header_row, c])
            if dept_name:
                dept_cols.append((c, dept_name))
        if not dept_cols:
            return None

        if not period:
            for ridx in range(min(8, len(raw))):
                row_text = "".join([_clean_text(v) for v in raw.iloc[ridx].tolist()])
                if "期间" in row_text:
                    period_guess = _extract_period_yyyymm(row_text)
                    if period_guess:
                        period = period_guess
                        break

        if not company:
            for ridx in range(max(0, len(raw) - 8), len(raw)):
                row_text = " ".join([str(v) for v in raw.iloc[ridx].tolist() if pd.notna(v)])
                if "单位:" in row_text:
                    company_part = row_text.split("单位:", 1)[-1]
                    company_part = company_part.split("操作员", 1)[0]
                    company_part = company_part.split("单位:", 1)[0]
                    company = company_part.strip()
                    break

        income_items = {"主营业务收入", "其他业务收入", "营业外收入"}
        cost_items = {"主营业务成本", "其他业务成本"}
        expense_items = {"销售费用", "管理费用", "财务费用"}

        agg: Dict[str, Dict[str, float]] = {}
        for _, dept in dept_cols:
            agg[dept] = {"income_amount": 0.0, "cost_amount": 0.0, "expense_amount": 0.0}

        start_row = header_row + 2
        for ridx in range(start_row, len(raw)):
            item_name = _clean_text(raw.iat[ridx, 0] if raw.shape[1] > 0 else "")
            if not item_name:
                continue
            if item_name.startswith("单位:"):
                break
            if item_name == "合计":
                break
            if item_name in {"会计科目", "统计方式", "余额方向"}:
                continue

            target_col = ""
            if item_name in income_items:
                target_col = "income_amount"
            elif item_name in cost_items:
                target_col = "cost_amount"
            elif item_name in expense_items:
                target_col = "expense_amount"
            else:
                continue

            for cidx, dept in dept_cols:
                agg[dept][target_col] += _to_float(raw.iat[ridx, cidx])

        rows = []
        for _, dept in dept_cols:
            income_amount = agg[dept]["income_amount"]
            cost_amount = agg[dept]["cost_amount"]
            expense_amount = agg[dept]["expense_amount"]
            rows.append({
                "company_code": company or "",
                "period": period or "",
                "dept_code": dept,
                "dept_name": dept,
                "income_amount": income_amount,
                "cost_amount": cost_amount,
                "expense_amount": expense_amount,
                "profit_amount": income_amount - cost_amount - expense_amount,
            })

        return pd.DataFrame(rows)

    def parse(self, file_path: str, **kwargs) -> pd.DataFrame:
        """解析管理中心部门收入成本费用表"""
        company = _sanitize_company_candidate(kwargs.get("company_code") or self.company_code or "")
        period = kwargs.get("period") or self.period or ""

        raw = self._read_excel(file_path, header=None)
        matrix_df = self._parse_matrix_layout(raw, company, period)
        if matrix_df is not None and len(matrix_df) > 0:
            return matrix_df

        df = self._read_excel(file_path)
        df = self._clean_df(df)
        df = df.rename(columns=self.COLUMN_MAPPING)

        df["company_code"] = company
        df["period"] = period

        for col in ["income_amount", "cost_amount", "expense_amount", "profit_amount"]:
            if col in df.columns:
                df[col] = self._normalize_numeric(df[col])

        # 计算利润
        if "profit_amount" not in df.columns:
            has_all = all(c in df.columns for c in ["income_amount", "cost_amount", "expense_amount"])
            if has_all:
                df["profit_amount"] = df["income_amount"] - df["cost_amount"] - df["expense_amount"]

        result_cols = ["company_code", "period", "dept_code", "dept_name",
                        "income_amount", "cost_amount", "expense_amount", "profit_amount"]
        available_cols = [c for c in result_cols if c in df.columns]
        return df[available_cols].copy()


# ============================================================================
# 非学科管理中心部门收入成本费用表解析器
# ============================================================================

class NonSubjectMgmtDeptIncomeCostParser(BaseParser):
    """非学科管理中心部门收入成本费用表解析器"""

    TABLE_NAME = "non_subject_mgmt_dept_income_cost"

    COLUMN_MAPPING = {
        "部门编码": "dept_code",
        "部门代码": "dept_code",
        "部门名称": "dept_name",
        "部门": "dept_name",
        "学科类型": "subject_type",
        "学科": "subject_type",
        "收入": "income_amount",
        "收入金额": "income_amount",
        "成本": "cost_amount",
        "成本金额": "cost_amount",
        "费用": "expense_amount",
        "费用金额": "expense_amount",
        "利润": "profit_amount",
        "利润金额": "profit_amount",
    }

    @staticmethod
    def _extract_dept_name(header_value: Any) -> str:
        text = _clean_text(header_value)
        if not text:
            return ""
        if "部门档案" in text:
            text = text.replace("【", "").replace("】", "")
            text = text.replace("部门档案:", "").replace("部门档案：", "")
            return _clean_text(text)
        return text

    def _parse_matrix_layout(self, raw: pd.DataFrame, company: str, period: str, subject_type: str) -> Optional[pd.DataFrame]:
        if not _looks_like_matrix_dept_report(raw):
            return None
        sample = raw.head(12).fillna("").astype(str)
        sample_text = "".join(sample.values.flatten().tolist())
        # 非学科矩阵通常带“素质管理中心/教研部/学术中心”特征。
        if not any(k in sample_text for k in ["素质管理中心", "素质中心教研部", "学术中心"]):
            return None

        header_row = None
        for ridx in range(min(12, len(raw))):
            row_values = [_clean_text(v) for v in raw.iloc[ridx].tolist()]
            row_text = "".join(row_values)
            if "会计科目" in row_text and any("部门档案" in x for x in row_values):
                header_row = ridx
                break
        if header_row is None:
            header_row = 5 if len(raw) > 7 else 0

        dept_cols: List[Tuple[int, str]] = []
        for c in range(3, raw.shape[1]):
            dept_name = self._extract_dept_name(raw.iat[header_row, c])
            if dept_name:
                dept_cols.append((c, dept_name))
        if not dept_cols:
            return None

        if not period:
            for ridx in range(min(8, len(raw))):
                row_text = "".join([_clean_text(v) for v in raw.iloc[ridx].tolist()])
                if "期间" in row_text:
                    period_guess = _extract_period_yyyymm(row_text)
                    if period_guess:
                        period = period_guess
                        break

        if not company:
            for ridx in range(max(0, len(raw) - 8), len(raw)):
                row_text = " ".join([str(v) for v in raw.iloc[ridx].tolist() if pd.notna(v)])
                if "单位:" in row_text:
                    company_part = row_text.split("单位:", 1)[-1]
                    company_part = company_part.split("操作员", 1)[0]
                    company_part = company_part.split("单位:", 1)[0]
                    company = company_part.strip()
                    break

        income_items = {"主营业务收入", "其他业务收入", "营业外收入"}
        cost_items = {"主营业务成本", "其他业务成本"}
        expense_items = {"销售费用", "管理费用", "财务费用"}

        agg: Dict[str, Dict[str, float]] = {}
        for _, dept in dept_cols:
            agg[dept] = {"income_amount": 0.0, "cost_amount": 0.0, "expense_amount": 0.0}

        start_row = header_row + 2
        for ridx in range(start_row, len(raw)):
            item_name = _clean_text(raw.iat[ridx, 0] if raw.shape[1] > 0 else "")
            if not item_name:
                continue
            if item_name.startswith("单位:"):
                break
            if item_name == "合计":
                break
            if item_name in {"会计科目", "统计方式", "余额方向"}:
                continue

            target_col = ""
            if item_name in income_items:
                target_col = "income_amount"
            elif item_name in cost_items:
                target_col = "cost_amount"
            elif item_name in expense_items:
                target_col = "expense_amount"
            else:
                continue

            for cidx, dept in dept_cols:
                agg[dept][target_col] += _to_float(raw.iat[ridx, cidx])

        rows = []
        for _, dept in dept_cols:
            income_amount = agg[dept]["income_amount"]
            cost_amount = agg[dept]["cost_amount"]
            expense_amount = agg[dept]["expense_amount"]
            rows.append({
                "company_code": company or "",
                "period": period or "",
                "dept_code": dept,
                "dept_name": dept,
                "subject_type": subject_type,
                "income_amount": income_amount,
                "cost_amount": cost_amount,
                "expense_amount": expense_amount,
                "profit_amount": income_amount - cost_amount - expense_amount,
            })

        return pd.DataFrame(rows)

    def parse(self, file_path: str, **kwargs) -> pd.DataFrame:
        """解析非学科管理中心部门收入成本费用表"""
        company = _sanitize_company_candidate(kwargs.get("company_code") or self.company_code or "")
        period = kwargs.get("period") or self.period or ""
        subject_type = kwargs.get("subject_type") or "非学科"

        raw = self._read_excel(file_path, header=None)
        matrix_df = self._parse_matrix_layout(raw, company, period, subject_type)
        if matrix_df is not None and len(matrix_df) > 0:
            return matrix_df

        df = self._read_excel(file_path)
        df = self._clean_df(df)
        df = df.rename(columns=self.COLUMN_MAPPING)

        df["company_code"] = company
        df["period"] = period
        if "subject_type" not in df.columns:
            df["subject_type"] = subject_type

        for col in ["income_amount", "cost_amount", "expense_amount", "profit_amount"]:
            if col in df.columns:
                df[col] = self._normalize_numeric(df[col])

        if "profit_amount" not in df.columns:
            has_all = all(c in df.columns for c in ["income_amount", "cost_amount", "expense_amount"])
            if has_all:
                df["profit_amount"] = df["income_amount"] - df["cost_amount"] - df["expense_amount"]

        result_cols = ["company_code", "period", "dept_code", "dept_name",
                        "subject_type", "income_amount", "cost_amount",
                        "expense_amount", "profit_amount"]
        available_cols = [c for c in result_cols if c in df.columns]
        return df[available_cols].copy()


# ============================================================================
# 非学科课酬表解析器
# ============================================================================

class NonSubjectTeachingFeeParser(BaseParser):
    """非学科课酬表解析器"""

    TABLE_NAME = "non_subject_teaching_fee"

    COLUMN_MAPPING = {
        "教师编码": "teacher_id",
        "教师代码": "teacher_id",
        "教师编号": "teacher_id",
        "职员代码": "teacher_id",
        "教师姓名": "teacher_name",
        "教师": "teacher_name",
        "职员姓名": "teacher_name",
        "课程类型": "course_type",
        "课程": "course_type",
        "部门名称": "source_dept_name",
        "打印部门名称": "course_type",
        "课时": "hours",
        "课时数": "hours",
        "授课时长": "hours",
        "单价": "rate",
        "课酬单价": "rate",
        "课酬": "total_amount",
        "课酬总额": "total_amount",
        "金额": "total_amount",
        "总收入": "total_amount",
    }

    def parse(self, file_path: str, **kwargs) -> pd.DataFrame:
        """解析非学科课酬表"""
        df = self._read_excel(file_path)
        df = self._clean_df(df)
        df = df.rename(columns=self.COLUMN_MAPPING)

        def _first_series(col_name: str) -> Optional[pd.Series]:
            if col_name not in df.columns:
                return None
            selected = df.loc[:, df.columns == col_name]
            if isinstance(selected, pd.DataFrame):
                return selected.bfill(axis=1).iloc[:, 0]
            return selected

        company = _sanitize_company_candidate(kwargs.get("company_code") or self.company_code or "")
        period = kwargs.get("period") or self.period or ""

        if not period:
            year_series = _first_series("年份")
            month_series = _first_series("月份")
            if year_series is not None and month_series is not None and len(df) > 0:
                year_non_null = year_series.dropna()
                month_non_null = month_series.dropna()
                year_val = str(year_non_null.iloc[0]) if not year_non_null.empty else ""
                month_val = str(month_non_null.iloc[0]) if not month_non_null.empty else ""
                period_guess = _extract_period_yyyymm(f"{year_val}-{month_val}")
                if period_guess:
                    period = period_guess
            if not period:
                period = _extract_period_yyyymm(Path(file_path).name)

        if not company:
            # 非学科课酬默认归口非学科管理中心，可在导入页手工覆盖。
            company = "东莞非学科管理中心"
            self.warnings.append("未识别到公司编码，已默认使用“东莞非学科管理中心”，可在导入设置手工覆盖。")

        df["company_code"] = company
        df["period"] = period

        has_total_amount = "total_amount" in df.columns
        for col in ["hours", "rate", "total_amount"]:
            if col in df.columns:
                series = _first_series(col)
                df[col] = self._normalize_numeric(series if series is not None else df[col])
            else:
                df[col] = 0.0

        if "course_type" not in df.columns:
            if "source_dept_name" in df.columns:
                df["course_type"] = df["source_dept_name"].astype(str).str.strip()
            else:
                df["course_type"] = "未分类"
        else:
            course_series = _first_series("course_type")
            if course_series is None:
                df["course_type"] = "未分类"
            else:
                df["course_type"] = course_series.astype(str).str.strip().replace({"": "未分类", "nan": "未分类"})

        if "teacher_id" in df.columns:
            teacher_id_series = _first_series("teacher_id")
            if teacher_id_series is not None:
                df["teacher_id"] = teacher_id_series.astype(str).str.strip()
        else:
            df["teacher_id"] = [f"ROW{i+1}" for i in range(len(df))]
        if "teacher_name" in df.columns:
            teacher_name_series = _first_series("teacher_name")
            if teacher_name_series is not None:
                df["teacher_name"] = teacher_name_series.astype(str).str.strip()
        else:
            df["teacher_name"] = ""

        # 计算课酬总额
        if (not has_total_amount) and "hours" in df.columns and "rate" in df.columns:
            df["total_amount"] = df["hours"] * df["rate"]

        result_cols = ["company_code", "period", "teacher_id", "teacher_name",
                        "course_type", "hours", "rate", "total_amount"]
        available_cols = [c for c in result_cols if c in df.columns]
        return df[available_cols].copy()


class BalanceSheetParser(BaseParser):
    """（合并）资产负债表解析器"""

    TABLE_NAME = "balance_sheet"

    def _read_sheet_cells(self, file_path: str):
        """读取 Excel 所有单元格，返回 (nrows, ncols, cell_value_func)"""
        ext = os.path.splitext(file_path)[1].lower()

        def _try_xlrd():
            import xlrd
            wb = xlrd.open_workbook(file_path)
            sh = wb.sheet_by_index(0)
            return sh.nrows, sh.ncols, lambda r, c: sh.cell_value(r, c)

        def _try_openpyxl():
            from openpyxl import load_workbook
            wb = load_workbook(file_path, data_only=True)
            ws = wb.worksheets[0]
            max_row = ws.max_row or 0
            max_col = ws.max_column or 0
            rows_data = []
            for row in ws.iter_rows(min_row=1, max_row=max_row, max_col=max_col, values_only=True):
                rows_data.append(list(row))
            return len(rows_data), max_col, lambda r, c: rows_data[r][c] if r < len(rows_data) and c < len(rows_data[r]) else ""

        # 先按扩展名尝试，失败则换另一种
        if ext == ".xls":
            try:
                return _try_xlrd()
            except Exception:
                try:
                    return _try_openpyxl()
                except Exception as e:
                    raise ValueError(f"无法读取 .xls 文件: {e}")
        else:
            # .xlsx 或未知扩展名，先试 openpyxl 再试 xlrd
            try:
                return _try_openpyxl()
            except Exception:
                try:
                    return _try_xlrd()
                except Exception as e:
                    raise ValueError(f"无法读取文件: {e}")

    def parse(self, file_path: str, **kwargs) -> pd.DataFrame:
        """解析资产负债表（支持 .xls 和 .xlsx）"""
        nrows, ncols, cell = self._read_sheet_cells(file_path)

        # Step 1: 找表头行
        header_row = None
        for r in range(min(nrows, 15)):
            row_vals = [str(cell(r, c)).strip() for c in range(ncols)]
            text = "".join(row_vals).replace(" ", "")
            if "资产" in text and "行次" in text and "期末余额" in text and "年初余额" in text:
                header_row = r
                break

        if header_row is None:
            self.errors.append("未找到资产负债表表头行")
            return pd.DataFrame()

        # Step 2: 提取公司和期间
        company = kwargs.get("company_code") or self.company_code or ""
        period = kwargs.get("period") or self.period or ""

        if not company or not period:
            for r in range(header_row - 1, max(header_row - 5, 0) - 1, -1):
                row_vals = [str(cell(r, c)).strip() for c in range(ncols)]
                full_text = " ".join(row_vals)
                if not company:
                    for v in row_vals:
                        if v.startswith("单位"):
                            m_comp = v
                            # "单位:拔创中心" 或 "单位:公司名"
                            parts = m_comp.replace("单位", "").replace("：", ":").split(":")
                            if len(parts) > 1 and parts[1].strip():
                                company = parts[1].strip()
                            break
                # 提取期间
                if not period:
                    for v in row_vals:
                        if "会计月" in v or "期间" in v:
                            m_per = v
                            for sep in ["：", ":", " "]:
                                if sep in m_per:
                                    parts = m_per.split(sep)
                                    if len(parts) > 1:
                                        raw = parts[-1].strip()
                                        # 提取 YYYYMM
                                        import re
                                        m = re.search(r"(\d{4})[年\-\s]*(\d{1,2})", raw)
                                        if m:
                                            period = m.group(1) + m.group(2).zfill(2)
                                        else:
                                            m2 = re.search(r"(\d{6})", raw)
                                            if m2:
                                                period = m2.group(1)

        # Step 3: 解析数据行
        rows = []
        section_names = [
            "流动资产：", "流动负债：", "非流动资产：", "非流动负债：",
            "所有者权益（或股东权益）：", "所有者权益：", "股东权益：",
            "流动资产", "流动负债", "非流动资产", "非流动负债",
            "所有者权益（或股东权益）"
        ]

        for r in range(header_row + 1, nrows):
            a = str(cell(r, 0)).strip()  # 资产名称
            b = str(cell(r, 1)).strip()  # 行次(左)
            c_val = cell(r, 2)            # 期末余额(左)
            d_val = cell(r, 3)            # 年初余额(左)
            e = str(cell(r, 4)).strip()  # 负债和所有者权益名称
            f = str(cell(r, 5)).strip()  # 行次(右)
            g_val = cell(r, 6)            # 期末余额(右)
            h_val = cell(r, 7)            # 年初余额(右)

            # 标准化数值
            def num(v):
                if isinstance(v, (int, float)):
                    return 0.0 if pd.isna(v) else float(v)
                return 0.0

            left_end = num(c_val)
            left_open = num(d_val)
            right_end = num(g_val)
            right_open = num(h_val)

            def row_type(name):
                if not name:
                    return -1
                n = name.replace(" ", "")
                if any(kw in n for kw in ["合计", "总计", "小计"]):
                    return 1
                for h in section_names:
                    if n == h.replace(" ", ""):
                        return 2
                return 0

            # 左项（资产侧）
            if a:
                rt = row_type(a)
                if rt >= 0:  # 所有非空都保留
                    sort_key = r - header_row
                    rows.append({
                        "company_code": company, "period": period,
                        "side": "资产", "item_name": a,
                        "line_number": b,
                        "ending_balance": left_end, "opening_balance": left_open,
                        "is_subtotal": rt, "sort_order": sort_key * 2,
                    })

            # 右项（负债和所有者权益侧）
            if e:
                rt = row_type(e)
                if rt >= 0:
                    rs = r - header_row
                    rows.append({
                        "company_code": company, "period": period,
                        "side": "负债和所有者权益", "item_name": e,
                        "line_number": f,
                        "ending_balance": right_end, "opening_balance": right_open,
                        "is_subtotal": rt, "sort_order": rs * 2 + 1,
                    })

        df_result = pd.DataFrame(rows)
        if len(df_result) == 0:
            self.errors.append("未解析到资产负债表数据行")
            return df_result

        return df_result


class IncomeStatementParser(BaseParser):
    """损益表解析器"""

    TABLE_NAME = "income_statement"

    def parse(self, file_path: str, **kwargs) -> pd.DataFrame:
        """解析损益表 Excel（支持单公司多期 和 多公司合集两种格式）"""
        # 支持 .xls 和 .xlsx 自动切换
        ext = os.path.splitext(file_path)[1].lower()
        ws = None
        if ext == ".xls":
            import xlrd
            wb = xlrd.open_workbook(file_path)
            sh = wb.sheet_by_index(0)
            # 包装成类似 openpyxl 的接口
            class XlrdSheet:
                def __init__(self, sh):
                    self._sh = sh
                    self.max_row = sh.nrows
                    self.max_column = sh.ncols
                def cell(self, r, c):
                    v = self._sh.cell_value(r-1, c-1)
                    return type('Cell', (), {'value': v})()
            ws = XlrdSheet(sh)
        else:
            try:
                from openpyxl import load_workbook
                wb = load_workbook(file_path, data_only=True)
                ws = wb.worksheets[0]
            except Exception:
                # openpyxl 失败则尝试 xlrd
                import xlrd
                wb = xlrd.open_workbook(file_path)
                sh = wb.sheet_by_index(0)
                class XlrdSheet:
                    def __init__(self, sh):
                        self._sh = sh
                        self.max_row = sh.nrows
                        self.max_column = sh.ncols
                    def cell(self, r, c):
                        v = self._sh.cell_value(r-1, c-1)
                        return type('Cell', (), {'value': v})()
                ws = XlrdSheet(sh)

        if ws is None:
            self.errors.append("无法读取文件")
            return pd.DataFrame()

        max_row = ws.max_row or 0
        max_col = ws.max_column or 0
        header_row = None
        for r in range(1, min(max_row, 10) + 1):
            v = str(ws.cell(r, 1).value or "").strip()
            if "项目" in v.replace(" ", ""):
                header_row = r
                break

        if header_row is None:
            self.errors.append("未找到损益表表头行")
            return pd.DataFrame()

        # 判断格式：表头第2列是否为中文公司名（多公司合集）还是日期/数字（单公司）
        second_col = str(ws.cell(header_row, 2).value or "").strip()
        is_multi_company = any("\u4e00" <= ch <= "\u9fff" for ch in second_col[:2])

        # 找标题行（含"项目"）
        header_row = None
        for r in range(1, min(max_row, 10) + 1):
            v = str(ws.cell(r, 1).value or "").strip()
            if "项目" in v.replace(" ", ""):
                header_row = r
                break

        if header_row is None:
            self.errors.append("未找到损益表表头行")
            return pd.DataFrame()

        # 判断格式：表头第2列是否为中文公司名（多公司合集）还是日期/数字（单公司）
        second_col = str(ws.cell(header_row, 2).value or "").strip()
        is_multi_company = any("\u4e00" <= ch <= "\u9fff" for ch in second_col[:2])

        # 提取公司和期间
        company = kwargs.get("company_code") or self.company_code or ""
        period = kwargs.get("period") or self.period or ""

        rows = []
        sort_idx = 0

        if is_multi_company:
            # === 多公司合集格式 ===
            # 构建名称→编码映射（从 database 读取，配置化）
            from .db_connection import execute_sql
            name_to_code = {}
            try:
                # 从 company_aliases 读简称映射
                alias_df = execute_sql("SELECT alias, company_code FROM company_aliases WHERE status = 1")
                for _, r in alias_df.iterrows():
                    name_to_code[r["alias"]] = r["company_code"]
                # 从 companies 读全称和简称
                comp_df = execute_sql("SELECT code, name, short_name FROM companies")
                for _, r in comp_df.iterrows():
                    name_to_code[r["name"]] = r["code"]
                    name_to_code[r["code"]] = r["code"]
                    if r["short_name"] and r["short_name"] != r["name"]:
                        name_to_code[r["short_name"]] = r["code"]
            except Exception:
                pass

            company_names_raw = []
            for c in range(2, max_col + 1):
                h = str(ws.cell(header_row, c).value or "").strip()
                if h:
                    company_names_raw.append(h)
                else:
                    break

            # 映射公司名→编码
            company_code_map = {}
            for cn in company_names_raw:
                if cn in name_to_code:
                    company_code_map[cn] = name_to_code[cn]
                else:
                    # 模糊匹配
                    matched = None
                    for full_name, code in sorted(name_to_code.items(), key=lambda x: -len(x[0])):
                        if isinstance(full_name, str) and (cn in full_name or full_name in cn):
                            if str(code).isdigit():
                                matched = code
                                break
                    company_code_map[cn] = matched or cn

            # 检测是否有重复公司名（模板可能列了两次：月度数 + 累计数）
            seen_names = set()
            unique_names = []
            for cn in company_names_raw:
                if cn not in seen_names:
                    seen_names.add(cn)
                    unique_names.append(cn)

            for r in range(header_row + 1, max_row + 1):
                item = str(ws.cell(r, 1).value or "").strip()
                if not item:
                    continue
                # 第一遍：读取各公司的月度值和累计值
                company_values = {}  # cname -> (monthly_val, cumulative_val)
                for ci, cname in enumerate(company_names_raw):
                    col = ci + 2
                    cell_val = ws.cell(r, col).value
                    try:
                        val = float(cell_val) if cell_val is not None else 0.0
                    except (ValueError, TypeError):
                        val = 0.0
                    if cname not in company_values:
                        company_values[cname] = [val, 0.0]  # [monthly, cumulative]
                    else:
                        company_values[cname][1] = val  # second occurrence = cumulative

                for ci, cname in enumerate(unique_names):
                    mapped_code = company_code_map.get(cname, cname)
                    mv, cv = company_values.get(cname, [0.0, 0.0])
                    rows.append({
                        "company_code": mapped_code,
                        "period": period,
                        "item_name": item,
                        "period1_value": mv,
                        "period2_value": 0.0,
                        "period3_value": 0.0,
                        "period4_value": 0.0,
                        "cumulative_value": cv,
                        "sort_order": sort_idx * 1000 + ci,
                        "original_name": cname,
                    })
                sort_idx += 1

            df_result = pd.DataFrame(rows)
            # 公司名映射将在 import_excel_to_db 中统一处理
            return df_result
        else:
            # === 单公司多期格式（月报/季报） ===
            import re as _re
            # 解析表头行的日期列，确定每列对应的 YYYYMM
            col_months = {}  # col_index -> YYYYMM
            for c in range(2, max_col + 1):
                h = str(ws.cell(header_row, c).value or "").strip()
                # 尝试匹配日期格式 "2026-01-01" 或 "2026年1月"
                m = _re.search(r"(\d{4})[年\-\s]*(\d{1,2})", h)
                if m:
                    col_months[c] = m.group(1) + m.group(2).zfill(2)

            # 找目标月份列：匹配导入期间（如 202603 → 找 202603 对应的列）
            target_col = None
            for c, ym in sorted(col_months.items()):
                if ym == period:
                    target_col = c
                    break
            # 没找到精确匹配则用第一列有数据的
            if target_col is None:
                target_col = 2

            # 找"本年累计"列
            cumul_col = None
            for c in range(2, max_col + 1):
                h = str(ws.cell(header_row, c).value or "").strip()
                if "累计" in h or "累积" in h:
                    cumul_col = c
                    break
            if cumul_col is None:
                cumul_col = max_col

            for r in range(header_row + 1, max_row + 1):
                item = str(ws.cell(r, 1).value or "").strip()
                if not item:
                    continue
                # 读取目标月份列的值
                main_val = 0.0
                cv = ws.cell(r, target_col).value
                try:
                    main_val = float(cv) if cv is not None else 0.0
                except (ValueError, TypeError):
                    main_val = 0.0

                # 读取累计值
                cumul_val = 0.0
                if cumul_col and cumul_col > 0:
                    cv2 = ws.cell(r, cumul_col).value
                    try:
                        cumul_val = float(cv2) if cv2 is not None else 0.0
                    except (ValueError, TypeError):
                        cumul_val = 0.0

                rows.append({
                    "company_code": company,
                    "period": period,
                    "item_name": item,
                    "period1_value": main_val,
                    "period2_value": cumul_val,
                    "period3_value": 0.0,
                    "period4_value": 0.0,
                    "cumulative_value": cumul_val,
                    "sort_order": sort_idx,
                })
                sort_idx += 1

            df_result = pd.DataFrame(rows)
            if len(df_result) == 0:
                self.errors.append("未解析到损益表数据行")
            return df_result


# ============================================================================
# 解析器注册表
# ============================================================================

# 报表类型 → 解析器类映射
PARSER_REGISTRY = {
    RT_ACCOUNT_BALANCE: AccountBalanceParser,
    RT_BALANCE_SHEET: BalanceSheetParser,
    RT_INCOME_STATEMENT: IncomeStatementParser,
    RT_PL_DETAIL: PlDetailParser,
    RT_INCOME_COST_EXPENSE: PlDetailParser,
    RT_REVENUE_VOLUME: RevenueVolumeParser,
    RT_NON_SUBJECT_ALLOCATION: NonSubjectAllocationParser,
    RT_MGMT_DEPT_INCOME_COST: MgmtDeptIncomeCostParser,
    RT_NON_SUBJECT_MGMT_DEPT_INCOME_COST: NonSubjectMgmtDeptIncomeCostParser,
    RT_NON_SUBJECT_TEACHING_FEE: NonSubjectTeachingFeeParser,
}


def get_parser(report_type_cn: str) -> Optional[BaseParser]:
    """
    根据报表中文类型名获取解析器实例

    Args:
        report_type_cn: 报表类型中文名

    Returns:
        解析器实例
    """
    parser_class = PARSER_REGISTRY.get(report_type_cn)
    if parser_class:
        return parser_class()
    return None


def parse_file(file_path: str, report_type: Optional[str] = None,
               company_code: Optional[str] = None,
               period: Optional[str] = None,
               original_filename: Optional[str] = None) -> Tuple[Optional[pd.DataFrame], str, Dict[str, Any]]:
    """
    解析单个 Excel 文件

    Args:
        file_path: Excel 文件路径
        report_type: 指定报表类型（如不指定则自动识别）
        company_code: 公司编码
        period: 期间 YYYYMM
        original_filename: 上传原始文件名（用于识别类型）

    Returns:
        (DataFrame, 报表类型, 解析结果信息)
    """
    result_info = {"errors": [], "warnings": []}

    # 检查文件是否存在
    if not os.path.exists(file_path):
        result_info["errors"].append(f"文件不存在: {file_path}")
        return None, "", result_info

    # 自动识别报表类型
    if not report_type:
        try:
            # 用 header=None 读取预览，避免矩阵模板被首行误判。
            df_preview = pd.read_excel(file_path, nrows=12, header=None)
            report_type = identify_report_type(file_path, df_preview, source_name=original_filename)
            if not report_type and original_filename:
                report_type = identify_report_type(original_filename, df_preview, source_name=original_filename)
        except Exception as e:
            result_info["errors"].append(f"无法读取文件: {e}")
            return None, "", result_info

    if not report_type:
        result_info["errors"].append(f"无法识别报表类型: {file_path}")
        return None, "", result_info

    # 获取解析器
    parser = get_parser(report_type)
    if not parser:
        result_info["errors"].append(f"不支持的报表类型: {report_type}")
        return None, report_type, result_info

    # 设置公司和期间
    if company_code:
        parser.company_code = company_code
    if period:
        parser.period = period

    # 解析
    try:
        df = parser.parse(file_path, company_code=company_code, period=period)
        result_info["errors"] = parser.errors
        result_info["warnings"] = parser.warnings

        if df is None or len(df) == 0:
            result_info["errors"].append("解析结果为空")
            return None, report_type, result_info

        result_info["rows"] = len(df)
        result_info["columns"] = list(df.columns)

        return df, report_type, result_info

    except Exception as e:
        result_info["errors"].append(f"解析失败: {e}")
        return None, report_type, result_info


def parse_and_validate(file_path: str, report_type: Optional[str] = None,
                        company_code: Optional[str] = None,
                        period: Optional[str] = None,
                        original_filename: Optional[str] = None):
    """
    解析并校验文件

    Args:
        file_path: Excel 文件路径
        report_type: 报表类型
        company_code: 公司编码
        period: 期间 YYYYMM

    Returns:
        (DataFrame, 报表类型, 校验结果, 解析信息)
    """
    from .validators import validate_import_file

    df, rtype, parse_info = parse_file(file_path, report_type, company_code, period, original_filename)

    if df is None:
        return None, rtype, parse_info, parse_info

    # 校验
    validation_result = validate_import_file(df, rtype)

    return df, rtype, validation_result, parse_info
