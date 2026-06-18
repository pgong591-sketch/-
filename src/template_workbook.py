"""Read-only access to fixed-format Excel report templates."""

from __future__ import annotations

from dataclasses import dataclass
from datetime import date, datetime
from html import escape
import os
from pathlib import Path
import re
from typing import Any
from xml.etree import ElementTree as ET

import openpyxl
import pandas as pd
from openpyxl.styles.colors import COLOR_INDEX
from openpyxl.utils import get_column_letter

from .db_connection import PROJECT_ROOT


TEMPLATE_WORKBOOK_NAME = "报表模板.xlsx"
PROJECT_TEMPLATE_PATH = PROJECT_ROOT / "data" / "templates" / TEMPLATE_WORKBOOK_NAME

DEFAULT_THEME_COLORS = [
    "000000",
    "FFFFFF",
    "1F497D",
    "EEECE1",
    "4F81BD",
    "C0504D",
    "9BBB59",
    "8064A2",
    "4BACC6",
    "F79646",
    "0000FF",
    "800080",
]

BORDER_WIDTHS = {
    "hair": "1px",
    "thin": "1px",
    "medium": "2px",
    "thick": "3px",
    "double": "3px",
    "dashed": "1px",
    "dotted": "1px",
    "mediumDashed": "2px",
    "dashDot": "1px",
    "mediumDashDot": "2px",
    "dashDotDot": "1px",
    "mediumDashDotDot": "2px",
    "slantDashDot": "2px",
}

BORDER_STYLES = {
    "double": "double",
    "dashed": "dashed",
    "mediumDashed": "dashed",
    "dotted": "dotted",
}


class TemplateWorkbookError(RuntimeError):
    """Raised when the fixed report template cannot be read."""


@dataclass(frozen=True)
class TemplateSheet:
    sheet_name: str
    source_path: Path
    row_count: int
    column_count: int
    html: str


def get_template_workbook_path() -> Path:
    """Return the fixed report template path without modifying the workbook."""
    env_path = os.environ.get("FINANCE_DW_REPORT_TEMPLATE_PATH")
    candidates: list[Path] = []
    if env_path:
        candidates.append(Path(env_path))

    candidates.append(PROJECT_TEMPLATE_PATH)

    desktop = Path.home() / "Desktop"
    if desktop.exists():
        candidates.extend(p for p in desktop.rglob("*.xlsx") if p.name == TEMPLATE_WORKBOOK_NAME)

    for path in candidates:
        if path.exists():
            return path

    raise TemplateWorkbookError(f"未找到报表模板文件：{TEMPLATE_WORKBOOK_NAME}")


def read_template_bytes() -> bytes:
    return get_template_workbook_path().read_bytes()


def load_template_sheet_frame(sheet_name: str) -> pd.DataFrame:
    """Load the exact sheet grid as rows and Excel column letters."""
    value_ws, formula_ws, path, _ = _open_sheet_pair(sheet_name)
    rows: list[list[Any]] = []
    for row_idx in range(1, value_ws.max_row + 1):
        row: list[Any] = []
        for col_idx in range(1, value_ws.max_column + 1):
            row.append(_cell_display_value(value_ws.cell(row_idx, col_idx), formula_ws.cell(row_idx, col_idx)))
        rows.append(row)

    columns = [get_column_letter(idx) for idx in range(1, value_ws.max_column + 1)]
    frame = pd.DataFrame(rows, columns=columns)
    frame.index = range(1, len(frame) + 1)
    frame.attrs["source_path"] = str(path)
    frame.attrs["sheet_name"] = sheet_name
    return frame


def load_template_sheet(
    sheet_name: str,
    min_col: int | None = None,
    max_col: int | None = None,
    min_row: int | None = None,
    max_row: int | None = None,
) -> TemplateSheet:
    """Render a fixed-format sheet as read-only HTML, preserving merges and basic styles."""
    value_ws, formula_ws, path, theme_colors = _open_sheet_pair(sheet_name)
    min_row, max_row, min_col, max_col = _normalize_bounds(value_ws, min_row, max_row, min_col, max_col)
    return TemplateSheet(
        sheet_name=sheet_name,
        source_path=path,
        row_count=max_row - min_row + 1,
        column_count=max_col - min_col + 1,
        html=_sheet_to_html(value_ws, formula_ws, min_row, max_row, min_col, max_col, theme_colors),
    )


def _open_sheet_pair(sheet_name: str):
    path = get_template_workbook_path()
    try:
        value_wb = openpyxl.load_workbook(path, data_only=True, read_only=False)
        formula_wb = openpyxl.load_workbook(path, data_only=False, read_only=False)
    except Exception as exc:
        raise TemplateWorkbookError(f"无法读取报表模板：{path}") from exc

    if sheet_name not in value_wb.sheetnames:
        raise TemplateWorkbookError(f"报表模板中不存在工作表：{sheet_name}")
    return value_wb[sheet_name], formula_wb[sheet_name], path, _extract_theme_colors(formula_wb)


def _cell_display_value(value_cell, formula_cell) -> Any:
    value = value_cell.value
    if value is None:
        value = formula_cell.value
    if value is None:
        return ""
    return value


def _normalize_bounds(ws, min_row, max_row, min_col, max_col) -> tuple[int, int, int, int]:
    min_row = max(1, min_row or 1)
    min_col = max(1, min_col or 1)
    max_row = min(ws.max_row, max_row or ws.max_row)
    max_col = min(ws.max_column, max_col or ws.max_column)
    if min_row > max_row or min_col > max_col:
        raise TemplateWorkbookError("模板显示范围无效")
    return min_row, max_row, min_col, max_col


def _visible_rows(ws, min_row: int, max_row: int) -> list[int]:
    return [
        row_idx
        for row_idx in range(min_row, max_row + 1)
        if not ws.row_dimensions[row_idx].hidden
    ]


def _visible_cols(ws, min_col: int, max_col: int) -> list[int]:
    return [
        col_idx
        for col_idx in range(min_col, max_col + 1)
        if not ws.column_dimensions[get_column_letter(col_idx)].hidden
    ]


def _sheet_to_html(value_ws, formula_ws, min_row: int, max_row: int, min_col: int, max_col: int, theme_colors: list[str]) -> str:
    row_indices = _visible_rows(value_ws, min_row, max_row)
    col_indices = _visible_cols(value_ws, min_col, max_col)

    merged_start: dict[tuple[int, int], tuple[int, int, int, int]] = {}
    merged_skip: set[tuple[int, int]] = set()
    for cell_range in value_ws.merged_cells.ranges:
        range_min_col, range_min_row, range_max_col, range_max_row = cell_range.bounds
        clipped_min_row = max(min_row, range_min_row)
        clipped_max_row = min(max_row, range_max_row)
        clipped_min_col = max(min_col, range_min_col)
        clipped_max_col = min(max_col, range_max_col)
        if clipped_min_row > clipped_max_row or clipped_min_col > clipped_max_col:
            continue

        display_rows = [idx for idx in row_indices if clipped_min_row <= idx <= clipped_max_row]
        display_cols = [idx for idx in col_indices if clipped_min_col <= idx <= clipped_max_col]
        if not display_rows or not display_cols:
            continue

        start = (display_rows[0], display_cols[0])
        merged_start[start] = (len(display_rows), len(display_cols), range_min_row, range_min_col)
        for row_idx in display_rows:
            for col_idx in display_cols:
                if (row_idx, col_idx) != start:
                    merged_skip.add((row_idx, col_idx))

    colgroup = "".join(
        f'<col style="width:{_column_width(value_ws, col_idx)}px">'
        for col_idx in col_indices
    )
    rows = ["<tbody>"]
    for row_idx in row_indices:
        height = value_ws.row_dimensions[row_idx].height
        row_style = f' style="height:{_points_to_px(height)}px"' if height else ""
        cells = [f"<tr{row_style}>"]
        for col_idx in col_indices:
            if (row_idx, col_idx) in merged_skip:
                continue

            rowspan, colspan, value_row, value_col = merged_start.get((row_idx, col_idx), (1, 1, row_idx, col_idx))
            value_cell = value_ws.cell(value_row, value_col)
            formula_cell = formula_ws.cell(value_row, value_col)
            span_attrs = ""
            if rowspan > 1:
                span_attrs += f' rowspan="{rowspan}"'
            if colspan > 1:
                span_attrs += f' colspan="{colspan}"'
            value = escape(_cell_display_text(value_cell, formula_cell))
            classes = []
            is_numeric = _is_numeric_display_cell(value_cell, formula_cell)
            if is_numeric:
                classes.append("numeric-cell")
            class_attr = f' class="{" ".join(classes)}"' if classes else ""
            cells.append(f'<td{class_attr}{span_attrs} style="{_cell_style(formula_cell, value_cell, theme_colors, is_numeric=is_numeric)}">{value}</td>')
        cells.append("</tr>")
        rows.append("".join(cells))
    rows.append("</tbody>")

    return (
        "<style>"
        ".template-sheet-wrap{overflow:auto;border:1px solid #d2d2d7;background:#f3f4f6;"
        "max-height:76vh;padding:12px;border-radius:8px;}"
        ".template-sheet{border-collapse:collapse;width:max-content;min-width:100%;background:#fff;"
        "font-family:'Microsoft YaHei','Noto Sans SC',Arial,sans-serif;color:#1d1d1f;"
        "box-shadow:0 1px 2px rgba(0,0,0,.05);}"
        ".template-sheet td{box-sizing:border-box;border:1px solid transparent;padding:5px 9px;"
        "min-width:52px;line-height:1.4;white-space:pre-wrap;overflow:visible;color:#111827;}"
        ".template-sheet td.numeric-cell{color:#111827!important;font-variant-numeric:tabular-nums;"
        "font-weight:650;text-shadow:none;opacity:1;}"
        "</style>"
        f'<div class="template-sheet-wrap"><table class="template-sheet"><colgroup>{colgroup}</colgroup>'
        + "".join(rows)
        + "</table></div>"
    )


def _column_width(ws, col_idx: int) -> int:
    letter = get_column_letter(col_idx)
    width = ws.column_dimensions[letter].width or 10
    return max(42, min(int((width + 0.75) * 7), 260))


def _points_to_px(points: float | None) -> int:
    if not points:
        return 22
    return max(18, int(points * 1.333))


def _cell_display_text(value_cell, formula_cell) -> str:
    value = _cell_display_value(value_cell, formula_cell)
    if value == "":
        return ""
    return str(_format_value(value, formula_cell))


def _is_numeric_display_cell(value_cell, formula_cell) -> bool:
    value = _cell_display_value(value_cell, formula_cell)
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return True
    text = str(_format_value(value, formula_cell)).strip() if value not in (None, "") else ""
    if not text or not any(char.isdigit() for char in text):
        return False
    normalized = text.replace(",", "").replace("，", "").replace(" ", "")
    return bool(re.fullmatch(r"[￥¥$€£()（）+\-–—\d.％%]+", normalized))


def _format_value(value: Any, cell) -> Any:
    if isinstance(value, str):
        return value
    number_format = cell.number_format or "General"
    if isinstance(value, datetime):
        if "年" in number_format and "月" in number_format and "日" not in number_format:
            return f"{value.year}年{value.month}月"
        if "年" in number_format and "月" in number_format:
            return f"{value.year}年{value.month}月{value.day}日"
        return value.strftime("%Y-%m-%d")
    if isinstance(value, date):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        if "%" in number_format:
            decimals = _decimal_places(number_format.split("%", 1)[0])
            return f"{value * 100:,.{decimals}f}%"
        if any(token in number_format for token in ("0", "#", "?")):
            decimals = _decimal_places(number_format)
            comma = "," in number_format
            if comma:
                return f"{value:,.{decimals}f}"
            return f"{value:.{decimals}f}"
        if isinstance(value, float) and value.is_integer():
            return int(value)
    return value


def _decimal_places(number_format: str) -> int:
    first_section = number_format.split(";", 1)[0]
    if "." not in first_section:
        return 0
    decimal_part = first_section.split(".", 1)[1]
    return sum(1 for char in decimal_part if char in "0#?")


def _cell_style(cell, value_cell, theme_colors: list[str], is_numeric: bool = False) -> str:
    style: list[str] = []
    background_color = None
    fill = cell.fill
    if fill and fill.fill_type and fill.fgColor:
        background_color = _color_to_hex(fill.fgColor, theme_colors)
        if background_color:
            style.append(f"background-color:{background_color}")

    font = cell.font
    if font:
        if font.bold:
            style.append("font-weight:700")
        if font.italic:
            style.append("font-style:italic")
        if font.sz:
            style.append(f"font-size:{font.sz}px")
        if font.name:
            style.append(f"font-family:'{font.name}','Microsoft YaHei','Noto Sans SC',Arial,sans-serif")
        font_color = _color_to_hex(font.color, theme_colors) if font.color else None
        safe_color = _readable_text_color(font_color, background_color, is_numeric=is_numeric)
        if safe_color:
            important = "!important" if is_numeric else ""
            style.append(f"color:{safe_color}{important}")
        if is_numeric and not font.bold:
            style.append("font-weight:650")
        if font.underline:
            style.append("text-decoration:underline")

    alignment = cell.alignment
    if alignment:
        horizontal = _css_horizontal(alignment.horizontal, value_cell.value)
        if horizontal:
            style.append(f"text-align:{horizontal}")
        if alignment.vertical:
            style.append(f"vertical-align:{alignment.vertical}")
        if alignment.wrap_text:
            style.append("white-space:pre-wrap")

    for edge in ("left", "right", "top", "bottom"):
        side = getattr(cell.border, edge)
        border = _border_style(side, edge, theme_colors)
        if border:
            style.append(border)

    return ";".join(style)


def _relative_luminance(hex_color: str) -> float:
    """Calculate WCAG relative luminance for a hex color."""
    color = str(hex_color or "").strip().lstrip("#")
    if len(color) == 3:
        color = "".join(ch * 2 for ch in color)
    if len(color) != 6:
        return 1.0

    channels = []
    for idx in (0, 2, 4):
        try:
            value = int(color[idx:idx + 2], 16) / 255
        except ValueError:
            value = 1.0
        if value <= 0.03928:
            channels.append(value / 12.92)
        else:
            channels.append(((value + 0.055) / 1.055) ** 2.4)
    return 0.2126 * channels[0] + 0.7152 * channels[1] + 0.0722 * channels[2]


def _contrast_ratio(foreground: str, background: str) -> float:
    """Calculate contrast ratio between foreground and background colors."""
    fg = _relative_luminance(foreground)
    bg = _relative_luminance(background)
    lighter = max(fg, bg)
    darker = min(fg, bg)
    return (lighter + 0.05) / (darker + 0.05)


def _readable_text_color(
    text_color: str | None,
    background_color: str | None,
    is_numeric: bool = False,
) -> str | None:
    """Return a readable text color when Excel colors have poor contrast."""
    background = background_color or "#ffffff"
    if is_numeric:
        return "#ffffff" if _relative_luminance(background) < 0.35 else "#111827"

    candidate = text_color or "#111827"
    if _contrast_ratio(candidate, background) >= 4.5:
        if text_color:
            return text_color
        if _relative_luminance(background) < 0.35:
            return "#ffffff"
        return "#111827" if is_numeric else None

    dark_text = "#111827"
    light_text = "#ffffff"
    if _relative_luminance(background) < 0.35:
        return light_text
    return dark_text


def _css_horizontal(horizontal: str | None, value: Any) -> str:
    if horizontal in {"center", "centerContinuous", "distributed", "justify"}:
        return "center"
    if horizontal in {"left", "right"}:
        return horizontal
    if isinstance(value, (int, float, datetime, date)) and not isinstance(value, bool):
        return "right"
    return "left"


def _border_style(side, edge: str, theme_colors: list[str]) -> str | None:
    if not side or not side.style:
        return None
    width = BORDER_WIDTHS.get(side.style, "1px")
    css_style = BORDER_STYLES.get(side.style, "solid")
    color = _color_to_hex(side.color, theme_colors) if side.color else "#9ca3af"
    return f"border-{edge}:{width} {css_style} {color}"


def _extract_theme_colors(workbook) -> list[str]:
    theme = getattr(workbook, "loaded_theme", None)
    if not theme:
        return DEFAULT_THEME_COLORS

    try:
        root = ET.fromstring(theme)
    except ET.ParseError:
        return DEFAULT_THEME_COLORS

    namespace = {"a": "http://schemas.openxmlformats.org/drawingml/2006/main"}
    scheme = root.find(".//a:clrScheme", namespace)
    if scheme is None:
        return DEFAULT_THEME_COLORS

    colors: list[str] = []
    for child in list(scheme):
        srgb = child.find(".//a:srgbClr", namespace)
        if srgb is not None and srgb.get("val"):
            colors.append(srgb.get("val", ""))
            continue
        system = child.find(".//a:sysClr", namespace)
        if system is not None and system.get("lastClr"):
            colors.append(system.get("lastClr", ""))

    return colors or DEFAULT_THEME_COLORS


def _color_to_hex(color, theme_colors: list[str]) -> str | None:
    if not color:
        return None
    if color.type == "rgb" and color.rgb:
        rgb = color.rgb[-6:]
        return f"#{rgb}"
    if color.type == "indexed" and isinstance(color.indexed, int) and color.indexed < len(COLOR_INDEX):
        return f"#{COLOR_INDEX[color.indexed][-6:]}"
    if color.type == "theme" and isinstance(color.theme, int):
        if color.theme < len(theme_colors):
            rgb = _apply_tint(theme_colors[color.theme], color.tint)
            return f"#{rgb}"
    return None


def _apply_tint(rgb: str, tint: float) -> str:
    rgb = rgb[-6:]
    if not tint:
        return rgb

    channels = [int(rgb[idx:idx + 2], 16) for idx in (0, 2, 4)]
    adjusted: list[int] = []
    for channel in channels:
        if tint < 0:
            value = int(channel * (1 + tint))
        else:
            value = int(channel + (255 - channel) * tint)
        adjusted.append(max(0, min(255, value)))
    return "".join(f"{channel:02X}" for channel in adjusted)
