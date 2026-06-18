from pathlib import Path

import pytest
from openpyxl import Workbook
from openpyxl.styles import Border, Font, PatternFill, Side

from src.template_workbook import (
    TemplateWorkbookError,
    _contrast_ratio,
    _readable_text_color,
    get_template_workbook_path,
    load_template_sheet,
    load_template_sheet_frame,
)


def _build_template(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "图片简报"
    ws["A1"] = "标题"
    ws["A1"].fill = PatternFill("solid", fgColor="1F4E78")
    ws["A1"].font = Font(color="FFFFFF", bold=True, size=16)
    ws["A2"] = "=1+1"
    ws["A3"] = "月报内容"
    ws["B3"] = "累计内容"
    ws["A4"] = 0.125
    ws["A4"].number_format = "0.0%"
    ws["B4"] = 1234.56
    ws["B4"].number_format = "#,##0.00"
    ws["A4"].border = Border(bottom=Side(style="thick", color="FF0000"))
    ws.merge_cells("A1:B1")

    income = wb.create_sheet("损益表")
    income["A1"] = "项目"
    income["B1"] = "本月数"
    income["A2"] = "营业收入"
    income["B2"] = 100

    summary = wb.create_sheet("经营汇总表")
    summary["A1"] = "项目"
    summary["B1"] = "金额"
    summary["A2"] = "净利润"
    summary["B2"] = 30
    summary["B2"].font = Font(color="FFFFFF")
    summary["B2"].fill = PatternFill("solid", fgColor="FFFFFF")
    summary["B3"] = "23.2%"
    summary["B3"].font = Font(color="FFFFFF")
    summary["B3"].fill = PatternFill("solid", fgColor="FFFFFF")
    wb.save(path)


def test_template_workbook_uses_configured_file_without_transforming_sheet(tmp_path, monkeypatch):
    template_path = tmp_path / "报表模板.xlsx"
    _build_template(template_path)
    monkeypatch.setenv("FINANCE_DW_REPORT_TEMPLATE_PATH", str(template_path))

    assert get_template_workbook_path() == template_path

    frame = load_template_sheet_frame("损益表")
    assert list(frame.columns) == ["A", "B"]
    assert frame.loc[1, "A"] == "项目"
    assert frame.loc[2, "A"] == "营业收入"
    assert frame.loc[2, "B"] == 100


def test_template_workbook_preview_preserves_merges_and_formula_text(tmp_path, monkeypatch):
    template_path = tmp_path / "报表模板.xlsx"
    _build_template(template_path)
    monkeypatch.setenv("FINANCE_DW_REPORT_TEMPLATE_PATH", str(template_path))

    sheet = load_template_sheet("图片简报")

    assert sheet.sheet_name == "图片简报"
    assert sheet.row_count == 4
    assert 'colspan="2"' in sheet.html
    assert "标题" in sheet.html
    assert "=1+1" in sheet.html
    assert "background-color:#1F4E78" in sheet.html
    assert "color:#FFFFFF" in sheet.html
    assert "font-weight:700" in sheet.html
    assert "12.5%" in sheet.html
    assert "1,234.56" in sheet.html
    assert 'class="numeric-cell"' in sheet.html
    assert "border-bottom:3px solid #FF0000" in sheet.html


def test_template_workbook_can_crop_picture_brief_columns(tmp_path, monkeypatch):
    template_path = tmp_path / "报表模板.xlsx"
    _build_template(template_path)
    monkeypatch.setenv("FINANCE_DW_REPORT_TEMPLATE_PATH", str(template_path))

    sheet = load_template_sheet("图片简报", min_col=1, max_col=1)

    assert "月报内容" in sheet.html
    assert "累计内容" not in sheet.html
    assert sheet.column_count == 1


def test_template_workbook_reports_missing_sheet(tmp_path, monkeypatch):
    template_path = tmp_path / "报表模板.xlsx"
    _build_template(template_path)
    monkeypatch.setenv("FINANCE_DW_REPORT_TEMPLATE_PATH", str(template_path))

    with pytest.raises(TemplateWorkbookError):
        load_template_sheet("不存在")


def test_readable_text_color_keeps_sufficient_contrast():
    assert _contrast_ratio("#111827", "#ffffff") >= 4.5
    assert _readable_text_color("#111827", "#ffffff") == "#111827"
    assert _readable_text_color("#ffffff", "#1F4E78") == "#ffffff"


def test_readable_text_color_fixes_low_contrast_on_light_background():
    assert _readable_text_color("#ffffff", "#ffffff", is_numeric=True) == "#111827"
    assert _readable_text_color("#ff0000", "#ffffff", is_numeric=True) == "#111827"
    assert _readable_text_color("#f8fafc", "#ffffff") == "#111827"


def test_readable_text_color_fixes_low_contrast_on_dark_background():
    assert _readable_text_color("#111827", "#000000") == "#ffffff"
    assert _readable_text_color(None, "#000000") == "#ffffff"


def test_numeric_template_cell_uses_readable_color(tmp_path, monkeypatch):
    template_path = tmp_path / "报表模板.xlsx"
    _build_template(template_path)
    monkeypatch.setenv("FINANCE_DW_REPORT_TEMPLATE_PATH", str(template_path))

    sheet = load_template_sheet("经营汇总表")

    assert "30" in sheet.html
    assert "23.2%" in sheet.html
    assert 'class="numeric-cell"' in sheet.html
    assert "color:#111827!important" in sheet.html
