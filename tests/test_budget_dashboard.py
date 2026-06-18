from pathlib import Path

import openpyxl

from src.budget_importer import CONFIRMED_BUDGET_SHEET, SUPPLEMENT_BUDGET_SHEET, import_budget_workbook
from src.dashboard_metrics import get_home_dashboard
from src.db_connection import get_connection, init_database


def _seed_company(conn):
    conn.execute(
        """
        INSERT INTO companies
            (code, name, short_name, parent_code, level, tree_path, is_leaf, is_consolidated, status)
        VALUES
            ('001', '莞城小学部', '莞城小学部', NULL, 0, '/001', 1, 1, 1)
        """
    )


def test_budget_importer_normalizes_targets_and_skips_subtotals(tmp_path):
    init_database()
    conn = get_connection()
    try:
        _seed_company(conn)
        conn.commit()
    finally:
        conn.close()

    workbook = openpyxl.Workbook()
    ws = workbook.active
    ws.title = CONFIRMED_BUDGET_SHEET
    ws.cell(3, 1, "素质中心")
    ws.cell(3, 2, "莞城小学部")
    ws.cell(3, 3, 1200)
    ws.cell(4, 2, "小计")
    ws.cell(4, 3, 9999)
    ws.cell(5, 2, "西平高中")
    ws.cell(5, 3, 8888)
    ws.cell(3, 21, "素质中心")
    ws.cell(3, 22, "莞城小学部")
    ws.cell(3, 23, 300)
    ws.cell(3, 24, 10)
    ws.cell(3, 25, 20)
    ws.cell(4, 22, "小计")
    ws.cell(4, 23, 9999)
    ws.cell(5, 22, "西平高中")
    ws.cell(5, 23, 8888)

    ws2 = workbook.create_sheet(SUPPLEMENT_BUDGET_SHEET)
    ws2.cell(3, 1, "素质中心")
    ws2.cell(3, 2, "莞城小学部")
    ws2.cell(3, 4, 100)
    ws2.cell(3, 5, 200)
    ws2.cell(3, 6, 300)
    ws2.cell(4, 2, "小计")
    ws2.cell(4, 4, 9999)
    ws2.cell(5, 2, "西平高中")
    ws2.cell(5, 4, 8888)

    path = tmp_path / "budget.xlsx"
    workbook.save(path)

    result = import_budget_workbook(str(path), budget_year="2026")
    assert result["success"] is True

    conn = get_connection()
    try:
        target_count = conn.execute("SELECT COUNT(*) FROM budget_targets").fetchone()[0]
        override_count = conn.execute("SELECT COUNT(*) FROM budget_actual_overrides").fetchone()[0]
        subtotal_count = conn.execute(
            "SELECT COUNT(*) FROM budget_targets WHERE project = '小计'"
        ).fetchone()[0]
        excluded_count = conn.execute(
            "SELECT COUNT(*) FROM budget_targets WHERE project = '西平高中'"
        ).fetchone()[0]
        company_code = conn.execute(
            "SELECT company_code FROM budget_targets WHERE project = '莞城小学部' LIMIT 1"
        ).fetchone()[0]
    finally:
        conn.close()

    assert target_count == 2
    assert override_count == 5
    assert subtotal_count == 0
    assert excluded_count == 0
    assert company_code == "001"


def test_home_dashboard_recalculates_ratios_and_uses_budget_overrides():
    init_database()
    conn = get_connection()
    try:
        _seed_company(conn)
        rows = [
            ("001", "202603", "一、营业收入", 300.0),
            ("001", "202603", "四、净利润（净亏损以“-”号填列）", 60.0),
            ("001", "202603", "净利率", 999.0),
            ("001", "202603", "减：营业成本", 120.0),
        ]
        conn.executemany(
            """
            INSERT INTO income_statement
                (company_code, period, item_name, period1_value, sort_order)
            VALUES (?, ?, ?, ?, 1)
            """,
            rows,
        )
        balance_rows = [
            ("001", "202603", "资产", "货币资金", 1000.0),
            ("001", "202603", "资产", "其他应收款", 100.0),
            ("001", "202603", "负债和所有者权益", "其他应付款", 50.0),
            ("001", "202603", "负债和所有者权益", "预收账款", 200.0),
            ("001", "202603", "资产", "资产总计", 1500.0),
            ("001", "202603", "负债和所有者权益", "负债和所有者权益（或股东权益）总计", 1500.0),
        ]
        conn.executemany(
            """
            INSERT INTO balance_sheet
                (company_code, period, side, item_name, ending_balance)
            VALUES (?, ?, ?, ?, ?)
            """,
            balance_rows,
        )
        conn.executemany(
            """
            INSERT INTO budget_targets
                (budget_year, module, project, company_code, target_type, annual_target)
            VALUES ('2026', '素质中心', '莞城小学部', '001', ?, ?)
            """,
            [("income", 1200.0), ("profit", 600.0)],
        )
        conn.executemany(
            """
            INSERT INTO budget_actual_overrides
                (budget_year, period, module, project, company_code, metric_type, amount)
            VALUES ('2026', ?, '素质中心', '莞城小学部', '001', ?, ?)
            """,
            [
                ("202601", "income", 100.0),
                ("202602", "income", 200.0),
                ("202601", "profit", 10.0),
                ("202602", "profit", 20.0),
            ],
        )
        conn.commit()
    finally:
        conn.close()

    dashboard = get_home_dashboard("202603", "001")

    assert dashboard["income"]["net_margin"] == 0.2
    assert dashboard["budget"]["income_actual_ytd"] == 600.0
    assert dashboard["budget"]["income_completion"] == 0.5
    assert dashboard["balance"]["balance_gap"] == 0
